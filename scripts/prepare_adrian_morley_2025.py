"""Adrian Morley Services Limited — YE 30/11/2025 draft pack from bank screenshots.

Source: iPhone email 17 Aug 2026 (siiphone@icloud.com) — 15 Lloyds Business
account 2660 monthly screenshots. Prior year pack YE 30/11/2024.
"""

from __future__ import annotations

import csv
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook as lx

CHART = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents"
    r"\Practice\Working Papers\Accology Chart.xlsx"
)
CLIENT = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Adrian Morley Services Limited"
)

NAVY = "1B365D"
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# date, description, signed amount (in +, out -), running balance after
TXNS = [
    (date(2024, 12, 2), "SPECTRUM RAIL LIMI", 2400.00, 3066.89),
    (date(2024, 12, 3), "LOAN - 01168340BBL", -173.41, 2893.48),
    (date(2024, 12, 3), "HFX HFX ECCLES C", -500.00, 2393.48),
    (date(2024, 12, 4), "HFX HFX ECCLES C", -500.00, 1893.48),
    (date(2024, 12, 5), "SPECTRUM RAIL LIMI", 1599.60, 3493.08),
    (date(2024, 12, 6), "ADRIAN MORLEY", -1660.00, 1833.08),
    (date(2024, 12, 6), "LOYD LOYD 10-12 HA", -500.00, 1333.08),
    (date(2024, 12, 17), "SAGE UK LTD", -39.60, 1293.48),
    (date(2024, 12, 17), "SERVICE CHARGES", -8.50, 1284.98),
    (date(2024, 12, 23), "IRIS PAYROLL SOLNS", -10.50, 1274.48),
    (date(2024, 12, 31), "SPECTRUM RAIL LIMI", 3999.60, 5274.08),
    (date(2025, 1, 2), "ADRIAN SANTANDER", -2000.00, 3274.08),
    (date(2025, 1, 2), "ADRIAN MORLEY", -1000.00, 2274.08),
    (date(2025, 1, 3), "LOAN - 01168340BBL", -173.05, 2101.03),
    (date(2025, 1, 8), "HMRC CUSTOMS AND E", -1319.87, 781.16),
    (date(2025, 1, 16), "SAGE UK LTD", -39.60, 741.56),
    (date(2025, 1, 20), "SERVICE CHARGES", -8.50, 733.06),
    (date(2025, 1, 31), "SPECTRUM RAIL LIMI", 1200.00, 1933.06),
    (date(2025, 1, 31), "HMRC - ACCOUNTS OF", -1036.97, 896.09),
    (date(2025, 2, 3), "LOAN - 01168340BBL", -172.29, 723.80),
    (date(2025, 2, 17), "SAGE UK LTD", -39.60, 684.20),
    (date(2025, 2, 18), "SERVICE CHARGES", -8.50, 675.70),
    (date(2025, 2, 24), "IRIS PAYROLL SOLNS", -11.58, 664.12),
    (date(2025, 2, 24), "IRIS PAYROLL SOLNS", -11.58, 652.54),
    (date(2025, 2, 28), "SPECTRUM RAIL LIMI", 1200.00, 1852.54),
    (date(2025, 2, 28), "ADRIAN MORLEY", -600.00, 1252.54),
    (date(2025, 3, 3), "LOAN - 01168340BBL", -171.81, 1080.73),
    (date(2025, 3, 17), "SAGE UK LTD", -56.40, 1024.33),
    (date(2025, 3, 18), "SERVICE CHARGES", -8.50, 1015.83),
    (date(2025, 3, 21), "ENVOLVE LIMITED", 1200.00, 2215.83),
    (date(2025, 3, 31), "SPECTRUM RAIL LIMI", 1200.00, 3415.83),
    (date(2025, 4, 3), "LOAN - 01168340BBL", -172.35, 3243.48),
    (date(2025, 4, 9), "DONE BROTHERS CAS", 360.00, 3603.48),
    (date(2025, 4, 17), "SAGE UK LTD", -56.40, 3547.08),
    (date(2025, 4, 22), "SERVICE CHARGES", -8.50, 3538.58),
    (date(2025, 4, 23), "IRIS PAYROLL SOLNS", -11.58, 3527.00),
    (date(2025, 4, 29), "HMRC CUSTOMS AND E", -455.40, 3071.60),
    (date(2025, 4, 30), "SPECTRUM RAIL LIMI", 1200.00, 4271.60),
    (date(2025, 5, 6), "LOAN - 01168340BBL", -171.46, 4100.14),
    (date(2025, 5, 15), "SAGE UK LTD", -56.40, 4043.74),
    (date(2025, 5, 16), "RL COMMERCIAL LIMI", 1200.00, 5243.74),
    (date(2025, 5, 19), "SERVICE CHARGES", -8.50, 5235.24),
    (date(2025, 5, 21), "IRIS PAYROLL SOLNS", -11.58, 5223.66),
    (date(2025, 5, 23), "EVOLUTION TECHNICA", 600.00, 5823.66),
    (date(2025, 5, 30), "SPECTRUM RAIL LIMI", 1200.00, 7023.66),
    (date(2025, 6, 3), "LOAN - 01168340BBL", -171.31, 6852.35),
    (date(2025, 6, 13), "DONE BROTHERS CAS", 720.00, 7572.35),
    (date(2025, 6, 17), "SAGE UK LTD", -56.40, 7515.95),
    (date(2025, 6, 17), "SERVICE CHARGES", -8.50, 7507.45),
    (date(2025, 6, 18), "IRIS PAYROLL SOLNS", -11.58, 7495.87),
    (date(2025, 6, 19), "ADRIAN SANTANDER", -7000.00, 495.87),
    (date(2025, 6, 27), "SPECTRUM RAIL LIMI", 1200.00, 1695.87),
    (date(2025, 7, 2), "HMRC CUSTOMS AND E", -871.20, 824.67),
    (date(2025, 7, 3), "LOAN - 01168340BBL", -170.78, 653.89),
    (date(2025, 7, 16), "RUGBY LEA NO2 AC", 4500.00, 5153.89),
    (date(2025, 7, 16), "HMRC - ACCOUNTS OF", -852.50, 4301.39),
    (date(2025, 7, 17), "SAGE UK LTD", -56.40, 4244.99),
    (date(2025, 7, 21), "IRIS PAYROLL SOLNS", -11.58, 4233.41),
    (date(2025, 7, 21), "IRIS PAYROLL SOLNS", -11.58, 4221.83),
    (date(2025, 7, 21), "SERVICE CHARGES", -8.50, 4213.33),
    (date(2025, 7, 22), "HMRC - ACCOUNTS OF", -555.24, 3658.09),
    (date(2025, 7, 23), "DONE BROTHERS CAS", 360.00, 4018.09),
    (date(2025, 7, 31), "SPECTRUM RAIL LIMI", 1200.00, 5218.09),
    (date(2025, 7, 31), "ADRIAN MORLEY", -4000.00, 1218.09),
    (date(2025, 8, 4), "HISPEC ELECTRICAL", 1200.00, 2418.09),
    (date(2025, 8, 4), "LOAN - 01168340BBL", -170.44, 2247.65),
    (date(2025, 8, 8), "DONE BROTHERS CAS", 360.00, 2607.65),
    (date(2025, 8, 15), "SAGE UK LTD", -56.40, 2551.25),
    (date(2025, 8, 18), "RUGBY LEA NO2 AC", 4500.00, 7051.25),
    (date(2025, 8, 19), "SERVICE CHARGES", -8.50, 7042.75),
    (date(2025, 8, 21), "IRIS PAYROLL SOLNS", -11.58, 7031.17),
    (date(2025, 8, 29), "SPECTRUM RAIL LIMI", 1200.00, 8231.17),
    (date(2025, 9, 1), "ADRIAN MORLEY", -4000.00, 4231.17),
    (date(2025, 9, 1), "RUGBY LEA NO2 AC", 211.50, 4442.67),
    (date(2025, 9, 3), "LOAN - 01168340BBL", -170.35, 4272.32),
    (date(2025, 9, 17), "DONE BROTHERS CAS", 360.00, 4632.32),
    (date(2025, 9, 17), "SAGE UK LTD", -56.40, 4575.92),
    (date(2025, 9, 17), "RUGBY LEA NO2 AC", 4500.00, 9075.92),
    (date(2025, 9, 19), "SERVICE CHARGES", -8.50, 9067.42),
    (date(2025, 9, 22), "IRIS PAYROLL SOLNS", -11.58, 9055.84),
    (date(2025, 9, 29), "ADRIAN MORLEY", -4000.00, 5055.84),
    (date(2025, 9, 29), "SPECTRUM RAIL LIMI", 1200.00, 6255.84),
    (date(2025, 10, 2), "EDMUNDSON ELEC 3YA", 1200.00, 7455.84),
    (date(2025, 10, 2), "HMRC CUSTOMS AND E", -3455.10, 4000.74),
    (date(2025, 10, 3), "LOAN - 01168340BBL", -169.75, 3830.99),
    (date(2025, 10, 16), "SAGE UK LTD", -56.40, 3774.59),
    (date(2025, 10, 20), "RUGBY LEA NO2 AC", 4500.00, 8274.59),
    (date(2025, 10, 20), "SERVICE CHARGES", -8.50, 8266.09),
    (date(2025, 10, 21), "IRIS PAYROLL SOLNS", -11.58, 8254.51),
    (date(2025, 10, 21), "ADRIAN MORLEY", -3000.00, 5254.51),
    (date(2025, 10, 22), "DONE BROTHERS CAS", 720.00, 5974.51),
    (date(2025, 10, 27), "ADRIAN SANTANDER", -3100.00, 2874.51),
    (date(2025, 10, 31), "ELECT RESOURCING L", 1200.00, 4074.51),
    (date(2025, 11, 3), "LOAN - 01168340BBL", -169.32, 3905.19),
    (date(2025, 11, 17), "RUGBY LEA NO2 AC", 4500.00, 8405.19),
    (date(2025, 11, 17), "SAGE UK LTD", -56.40, 8348.79),
    (date(2025, 11, 18), "SERVICE CHARGES", -8.50, 8340.29),
    (date(2025, 11, 26), "ADRIAN MORLEY", -1400.00, 6940.29),
    (date(2025, 11, 28), "ELECT RESOURCING L", 1200.00, 8140.29),
]

SALES_PAYEES = {
    "SPECTRUM RAIL LIMI",
    "DONE BROTHERS CAS",
    "RUGBY LEA NO2 AC",
    "ELECT RESOURCING L",
    "EDMUNDSON ELEC 3YA",
    "HISPEC ELECTRICAL",
    "ENVOLVE LIMITED",
    "RL COMMERCIAL LIMI",
    "EVOLUTION TECHNICA",
}


def chart_names() -> set[str]:
    wb = lx(CHART, data_only=True)
    names = {str(r[0]).strip() for r in wb.active.iter_rows(min_row=2, values_only=True) if r and r[0]}
    wb.close()
    return names


def vat_split(gross: float) -> tuple[float, float]:
    net = round(abs(gross) / 1.2, 2)
    vat = round(abs(gross) - net, 2)
    return net, vat


def classify(desc: str, amt: float) -> tuple[str, float, float]:
    """Return (iris_or_tag, p_and_l_or_bs_net, vat_to_control). Signs: + = debit."""
    if amt > 0 and desc in SALES_PAYEES:
        net, vat = vat_split(amt)
        return "Turnover - Sales", -net, -vat
    if desc == "SERVICE CHARGES":
        return "Administrative expenses - General - Bank charges", -amt, 0.0
    if desc.startswith("SAGE") or desc.startswith("IRIS PAYROLL"):
        net, vat = vat_split(-amt)
        return "Administrative expenses - General - Software", net, vat
    if desc.startswith("HMRC CUSTOMS"):
        return "vat_paid", 0.0, -amt
    if desc.startswith("HMRC - ACCOUNTS"):
        return "Creditors less than 1 year - Other taxes and social security", -amt, 0.0
    if desc.startswith("LOAN -") or desc.startswith("ADRIAN ") or desc.startswith("HFX") or desc.startswith("LOYD"):
        return "dla", -amt, 0.0
    return "Administrative expenses - General - Sundry expenses", -amt, 0.0


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(color="FFFFFF", bold=True, name="Calibri")
    for c in range(1, cols + 1):
        ws.cell(row, c).fill = fill
        ws.cell(row, c).font = font


def write_money(ws, r, c, v):
    cell = ws.cell(r, c, round(float(v or 0), 2))
    cell.number_format = '#,##0.00;(#,##0.00);"—"'
    cell.alignment = Alignment(horizontal="right")


def write_iris(path: Path, buckets: dict[str, float], names: set[str]) -> tuple[float, list[str]]:
    unknown = [a for a in buckets if a not in names and abs(buckets[a]) >= 0.005]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Account", "Description", "Year End 30/11/2025"])
        for acc in sorted(buckets):
            amt = round(buckets[acc], 2)
            if abs(amt) < 0.005:
                continue
            w.writerow([acc, acc, f"{amt:.2f}"])
    return round(sum(buckets.values()), 2), unknown


def main() -> int:
    names = chart_names()
    # Opening 1 Dec 2024 from last year's last bank line
    open_bank = 666.89
    open_vat = -659.93  # credit
    open_acc = -1000.00
    open_sc = -100.00
    open_dla = 0.0
    open_re = round(-(open_bank + open_vat + open_acc + open_sc + open_dla), 2)

    buckets: dict[str, float] = defaultdict(float)
    buckets["Cash - Cash at bank and in hand"] = open_bank
    buckets["Creditors less than 1 year - Other taxes and social security"] = open_vat
    buckets["Creditors less than 1 year - Accruals"] = open_acc
    buckets["Share capital - Brought forward"] = open_sc
    buckets["Profit and loss account - Brought forward"] = open_re
    buckets["Creditors less than 1 year - Directors' loans"] = open_dla

    led_rows = []
    sales_by: dict[str, float] = defaultdict(float)
    for d, desc, amt, bal in TXNS:
        tag, pl, vat = classify(desc, amt)
        buckets["Cash - Cash at bank and in hand"] += amt
        if tag == "dla":
            buckets["Creditors less than 1 year - Directors' loans"] += pl
        elif tag == "vat_paid":
            buckets["Creditors less than 1 year - Other taxes and social security"] += vat
        else:
            buckets[tag] += pl
            if abs(vat) >= 0.005:
                buckets["Creditors less than 1 year - Other taxes and social security"] += vat
        if tag == "Turnover - Sales":
            sales_by[desc] += -pl
        led_rows.append((d, desc, amt, bal, tag, pl, vat))

    # Year-end journals — same policy as 2024 pack
    jnls = [
        ("Jnl1", "Administrative expenses - Employee costs - Directors' salaries", 23700.00, "Payroll 12 x £987.50 x 2 (same as 2024)"),
        ("Jnl1", "Creditors less than 1 year - Directors' loans", -23700.00, "Salary via DLA"),
        ("Jnl2", "Administrative expenses - Employee costs - Employer's NI", 865.93, "Draft — same as 2024 pending RTI"),
        ("Jnl2", "Creditors less than 1 year - Directors' loans", -865.93, "ENI via DLA"),
        ("Jnl3", "Administrative expenses - Employee costs - Travel and subsistence", 5000.00, "Mileage provision — same as 2024, confirm"),
        ("Jnl3", "Creditors less than 1 year - Directors' loans", -5000.00, "Mileage via DLA"),
        ("Jnl4", "Creditors less than 1 year - Accruals", 1000.00, "Release 2024 accountancy accrual"),
        ("Jnl4", "Creditors less than 1 year - Directors' loans", -1000.00, "Assume 2024 fee paid personally"),
        ("Jnl5", "Administrative expenses - Legal & professional - Accountancy fees", 1200.00, "CRM Accounts 2025-11-30 fee"),
        ("Jnl5", "Creditors less than 1 year - Accruals", -1200.00, "2025 accountancy accrual"),
    ]
    for _j, acc, amt, _n in jnls:
        buckets[acc] += amt

    for k in list(buckets):
        buckets[k] = round(buckets[k], 2)

    sales = -buckets.get("Turnover - Sales", 0)
    bank_ch = buckets.get("Administrative expenses - General - Bank charges", 0)
    software = buckets.get("Administrative expenses - General - Software", 0)
    dirsals = buckets.get("Administrative expenses - Employee costs - Directors' salaries", 0)
    eni = buckets.get("Administrative expenses - Employee costs - Employer's NI", 0)
    travel = buckets.get("Administrative expenses - Employee costs - Travel and subsistence", 0)
    accy = buckets.get("Administrative expenses - Legal & professional - Accountancy fees", 0)
    overheads = round(bank_ch + software + dirsals + eni + travel + accy, 2)
    profit = round(sales - overheads, 2)

    bank = buckets.get("Cash - Cash at bank and in hand", 0)
    vat = -buckets.get("Creditors less than 1 year - Other taxes and social security", 0)
    accruals = -buckets.get("Creditors less than 1 year - Accruals", 0)
    dla = -buckets.get("Creditors less than 1 year - Directors' loans", 0)
    sc = -buckets.get("Share capital - Brought forward", 0)
    re = -buckets.get("Profit and loss account - Brought forward", 0)
    net_assets = round(bank - vat - accruals - dla, 2)
    funds = round(sc + re + profit, 2)

    queries = [
        "Source is 15 Lloyds Business 2660 monthly screenshots emailed from siiphone@icloud.com on 17 Aug 2026. Not a full statement CSV. Small items could be off-screen if a month had more lines than the phone showed.",
        "Companies House final reminder: accounts 1 Dec 2024 to 30 Nov 2025 are due 31 August 2026.",
        "2024 working papers left a bank difference on the balance sheet (TB bank -£3,173). Actual Lloyds balance at 30 Nov 2024 was £666.89 (agrees to first Dec 2024 screenshot). Draft opens on the real bank, not the 2024 TB difference.",
        "All receipts treated as standard-rated VAT-inclusive (20%), same as 2024 Sage invoices (e.g. Spectrum £3,999.60 = £3,333 + VAT). Confirm Rugby Lea £4,500, Elect Resourcing, Hispec, Envolve, Edmundson, RL Commercial, Evolution Technica are all standard-rated sales.",
        "No Sage/sales invoice listing for 2025. Last year sales were invoiced in Sage (Impact/Frameworks/Betfred). Bank payees have changed (Rugby Lea, Elect, Spectrum still). Need sales invoices / Sage TB if any unpaid debtors exist — draft is cash receipts only.",
        "Directors' salaries £23,700 and employer's NI £865.93 copied from 2024 (payroll £987.50pm each). Confirm RTI / FPS and whether Clare is still on payroll.",
        "Mileage £5,000 copied from 2024 DLA journal. Confirm 2025 claim.",
        "HMRC Customs payments (£1,319.87 + £455.40 + £871.20 + £3,455.10 = £6,101.57) treated as VAT paid. October £3,455 is large — check VAT returns vs output VAT on this pack.",
        "HMRC Accounts Office payments (£1,036.97 + £852.50 + £555.24 = £2,444.71) treated as PAYE/NIC creditor payments. Confirm vs FPS. Not corporation tax (Accounts Office is usually PAYE).",
        "Bounce Back / LOAN 01168340BBL monthly ~£170 posted to DLA (same as 2024). Confirm remaining balance and whether any interest should hit P&L.",
        "Large drawings: June Santander £7,000; Sep two lots of £4,000; Oct Morley £3,000 + Santander £3,100. Dividend vs DLA vs salary — draft leaves them on DLA. 2024 declared a £10,000 dividend via DLA. Decide 2025 dividend before finals.",
        "2024 accountancy accrual £1,000 released to DLA (assumed paid personally). 2025 accrued at CRM fee £1,200.",
        "No stock, FA or HP on last year's BS. None identified this year.",
        "Corporation tax not computed. Draft profit is before tax. 2024 profit was £4,438 and CT was cleared.",
        "Share capital £100 brought forward. Confirm still 100 £1 shares.",
        "VAT number not on the CRM client record. Last year there was a VAT control. Confirm still VAT registered.",
    ]

    dest_pack = CLIENT / "Current" / "Working Papers" / "Adrian Morley Services Limited - 30 November 2025 - Draft accounts pack.xlsx"
    dest_iris = CLIENT / "Current" / "IRIS Import" / "2025-11-30 IRIS Elements TB.csv"
    dest_q = CLIENT / "Current" / "Working Papers" / "Adrian Morley Services Limited - 30 November 2025 - Queries.md"
    dest_csv = CLIENT / "Current" / "Source" / "2025-11-30 bank ledger from screenshots.csv"

    dest_csv.parent.mkdir(parents=True, exist_ok=True)
    with dest_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Date", "Description", "Amount", "Balance", "IRIS / treatment", "P&L/BS net", "VAT"])
        w.writerow(["2024-11-30", "Opening bank per 2024 last line", open_bank, open_bank, "Cash", "", ""])
        for row in led_rows:
            w.writerow([row[0].isoformat(), row[1], f"{row[2]:.2f}", f"{row[3]:.2f}", row[4], f"{row[5]:.2f}", f"{row[6]:.2f}"])

    net, unknown = write_iris(dest_iris, buckets, names)
    dest_q.write_text(
        "# Adrian Morley Services Limited — draft queries\n\nYear ended 30 November 2025. Draft only.\n\n"
        + "\n".join(f"{i}. {q}" for i, q in enumerate(queries, 1))
        + "\n",
        encoding="utf-8",
    )

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Adrian Morley Services Limited"
    cover["A1"].font = Font(name="Calibri", bold=True, size=18, color=NAVY)
    cover["A2"] = "Draft accounts working pack — year ended 30 November 2025"
    cover["A3"] = (
        "Prepared from Lloyds Business 2660 monthly screenshots (iPhone email 17 Aug 2026) "
        "and the 30 November 2024 working papers. Company 09888266."
    )
    cover["A4"] = (
        "NOT an IRIS statutory set. CH filing deadline 31 August 2026. "
        "Confirm queries before import / filing."
    )
    cover["A4"].font = Font(italic=True, color="833C0C")
    stats = [
        ("Opening bank 30 Nov 2024 (actual)", open_bank),
        ("Closing bank 30 Nov 2025", bank),
        ("Turnover (VAT-exclusive)", sales),
        ("Draft profit before tax", profit),
        ("VAT creditor / (debtor)", vat),
        ("Directors' loan (credit = company owes)", dla),
        ("Net assets / funds", net_assets),
        ("IRIS TB net", net),
        ("Comparatives: 2024 turnover", 35920.39),
        ("Comparatives: 2024 PBT", 4438.48),
    ]
    cover["A6"] = "Draft figures"
    cover["A6"].font = Font(bold=True, color=NAVY)
    for i, (lab, val) in enumerate(stats, 7):
        cover.cell(i, 1, lab)
        write_money(cover, i, 2, val)
    cover["A18"] = "Open queries"
    cover["A18"].font = Font(bold=True, color=NAVY)
    for i, q in enumerate(queries, 19):
        cover.cell(i, 1, f"{i-18}. {q}")
        cover.cell(i, 1).alignment = Alignment(wrap_text=True)
        cover.row_dimensions[i].height = 28
    cover.column_dimensions["A"].width = 112
    cover.column_dimensions["B"].width = 16

    qws = wb.create_sheet("Queries")
    qws["A1"] = "No"
    qws["B1"] = "Query"
    style_header(qws, 1, 2)
    for i, q in enumerate(queries, 1):
        qws.cell(i + 1, 1, i)
        qws.cell(i + 1, 2, q)
        qws.cell(i + 1, 2).alignment = Alignment(wrap_text=True, vertical="top")
        qws.row_dimensions[i + 1].height = 32
    qws.column_dimensions["A"].width = 6
    qws.column_dimensions["B"].width = 112

    led = wb.create_sheet("Bank ledger")
    for col, h in enumerate(["Date", "Description", "In", "Out", "Balance", "Treatment"], 1):
        led.cell(1, col, h)
    style_header(led, 1, 6)
    r = 2
    for d, desc, amt, bal, tag, pl, vat in led_rows:
        led.cell(r, 1, d.isoformat())
        led.cell(r, 2, desc)
        if amt > 0:
            write_money(led, r, 3, amt)
        else:
            write_money(led, r, 4, -amt)
        write_money(led, r, 5, bal)
        led.cell(r, 6, tag)
        r += 1
    for col, w in (("A", 14), ("B", 28), ("C", 12), ("D", 12), ("E", 12), ("F", 62)):
        led.column_dimensions[col].width = w

    jws = wb.create_sheet("Draft journals")
    jws["A1"] = "Jnl"
    jws["B1"] = "Account"
    jws["C1"] = "Debit"
    jws["D1"] = "Credit"
    jws["E1"] = "Narrative"
    style_header(jws, 1, 5)
    for i, (j, acc, amt, narr) in enumerate(jnls, 2):
        jws.cell(i, 1, j)
        jws.cell(i, 2, acc)
        if amt >= 0:
            write_money(jws, i, 3, amt)
        else:
            write_money(jws, i, 4, -amt)
        jws.cell(i, 5, narr)
    for col, w in (("A", 8), ("B", 62), ("C", 12), ("D", 12), ("E", 56)):
        jws.column_dimensions[col].width = w

    pl = wb.create_sheet("Draft P&L")
    pl["A1"] = "Profit and loss — year ended 30 November 2025"
    pl["A1"].font = Font(bold=True, size=14, color=NAVY)
    pl["A3"] = "Line"
    pl["B3"] = "2025 £"
    pl["C3"] = "2024 £"
    style_header(pl, 3, 3)
    pl_lines = [
        ("Turnover", sales, 35920.39),
        ("Directors' salaries", -dirsals, -23700),
        ("Employer's NI (draft)", -eni, -865.93),
        ("Travel / mileage (draft)", -travel, -5000),
        ("Bank charges", -bank_ch, -87),
        ("Software (Sage / IRIS payroll)", -software, -542.98),
        ("Repairs", 0.0, -286),
        ("Accountancy", -accy, -1000),
        ("Profit before tax", profit, 4438.48),
        ("Tax (not computed)", 0.0, 0.0),
        ("Profit for the year (before tax)", profit, 4438.48),
    ]
    for i, (lab, a, b) in enumerate(pl_lines, 4):
        pl.cell(i, 1, lab)
        write_money(pl, i, 2, a)
        write_money(pl, i, 3, b)
        if "Profit" in lab or lab == "Turnover":
            pl.cell(i, 1).font = Font(bold=True)
    r = 16
    pl.cell(r, 1, "Turnover by payee (net)")
    pl.cell(r, 1).font = Font(bold=True, color=NAVY)
    r += 1
    for name, netv in sorted(sales_by.items(), key=lambda x: -x[1]):
        pl.cell(r, 1, name)
        write_money(pl, r, 2, netv)
        r += 1
    pl.column_dimensions["A"].width = 42
    pl.column_dimensions["B"].width = 14
    pl.column_dimensions["C"].width = 14

    bs = wb.create_sheet("Draft balance sheet")
    bs["A1"] = "Balance sheet — 30 November 2025"
    bs["A1"].font = Font(bold=True, size=14, color=NAVY)
    bs["A3"] = "Line"
    bs["B3"] = "2025 £"
    style_header(bs, 3, 2)
    bs_lines = [
        ("Cash at bank", bank),
        ("VAT / PAYE (net creditor)", -vat),
        ("Accruals", -accruals),
        ("Directors' loan", -dla),
        ("Net current assets / (liabilities)", net_assets),
        ("Called up share capital", sc),
        ("Profit and loss brought forward", re),
        ("Profit for the year (before tax)", profit),
        ("Shareholders' funds", funds),
    ]
    for i, (lab, v) in enumerate(bs_lines, 4):
        bs.cell(i, 1, lab)
        write_money(bs, i, 2, v)
        if "Net" in lab or "funds" in lab:
            bs.cell(i, 1).font = Font(bold=True)
    bs.column_dimensions["A"].width = 42
    bs.column_dimensions["B"].width = 16

    iris = wb.create_sheet("IRIS Elements TB")
    iris["A1"] = "Account"
    iris["B1"] = "Description"
    iris["C1"] = "Year ended 30 November 2025"
    style_header(iris, 1, 3)
    r = 2
    tot = 0.0
    for acc in sorted(buckets):
        amt = round(buckets[acc], 2)
        if abs(amt) < 0.005:
            continue
        iris.cell(r, 1, acc)
        iris.cell(r, 2, acc)
        write_money(iris, r, 3, amt)
        tot += amt
        r += 1
    iris.cell(r, 1, "Net (must be 0.00)")
    iris.cell(r, 1).font = Font(bold=True)
    write_money(iris, r, 3, tot)
    iris.column_dimensions["A"].width = 62
    iris.column_dimensions["B"].width = 62
    iris.column_dimensions["C"].width = 22

    dest_pack.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest_pack)
    print(f"pack={dest_pack}")
    print(f"iris={dest_iris} net={net} unknown={unknown}")
    print(f"sales={sales} profit={profit} bank={bank} vat={vat} dla={dla}")
    print(f"na={net_assets} funds={funds} open_re={open_re}")
    print(f"ledger={dest_csv} n={len(TXNS)}")
    if unknown:
        print("UNMAPPED", unknown)
        return 1
    if abs(net) > 0.05:
        print("TB OUT")
        return 1
    if abs(net_assets - funds) > 0.05:
        print("BS mismatch")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
