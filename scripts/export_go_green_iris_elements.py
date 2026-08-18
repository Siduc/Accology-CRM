"""Map Go Green Mike TB onto Accology Chart (IRIS Elements / Taxfiler names)."""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

CHART = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents"
    r"\Practice\Working Papers\Accology Chart.xlsx"
)
SRC = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Go Green Pelleting Solutions Limited\Current\Source"
    r"\Go-Green-Pelleting-Solutions-2026-accounts.xlsx"
)
OUT = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Go Green Pelleting Solutions Limited\Current\IRIS Import"
    r"\2026-03-31 IRIS Elements TB.csv"
)
MAP_OUT = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Go Green Pelleting Solutions Limited\Current\IRIS Import"
    r"\2026-03-31 Accology chart mapping.csv"
)

# Mike line (or sheet row) -> Accology Chart IRIS name
BY_ROW = {
    11: "Plant & machinery - Cost - b/fwd",
    12: "Plant & machinery - Depn - b/fwd",
    13: "Plant & machinery - Cost - additions",
    14: "Plant & machinery - Cost - additions",  # transfer onto owned
    15: "Plant & machinery - Depn - b/fwd",  # depn coming across with the asset
    16: "Plant & machinery - Depn - charge for the year",
    17: "Plant & machinery - Cost - b/fwd",  # leased b/f
    18: "Plant & machinery - Depn - b/fwd",
    19: "Plant & machinery - Cost - disposals",  # lease transferred off
    20: "Plant & machinery - Depn - disposals",
    21: "Plant & machinery - Cost - additions",  # new HP/lease plant
    22: "Plant & machinery - Depn - charge for the year",
    23: "Fixtures & fittings - Cost - b/fwd",
    24: "Fixtures & fittings - Depn - b/fwd",
    25: "Fixtures & fittings - Depn - charge for the year",
    26: "Motor vehicles - Cost - additions",  # transfer to ownership
    27: "Motor vehicles - Cost - disposals",
    28: "Motor vehicles - Depn - b/fwd",
    29: "Motor vehicles - Depn - disposals",
    30: "Motor vehicles - Depn - charge for the year",
    31: "Motor vehicles - Cost - b/fwd",
    32: "Motor vehicles - Cost - disposals",  # lease transferred to owned
    33: "Motor vehicles - Depn - b/fwd",
    34: "Motor vehicles - Depn - disposals",
    35: "Motor vehicles - Depn - charge for the year",
    36: "Computer equipment - Cost - b/fwd",
    37: "Computer equipment - Depn - b/fwd",
    38: "Computer equipment - Depn - charge for the year",
    39: "Debtors - Other debtors",
    40: "Cash - Cash at bank and in hand",
    41: "Creditors less than 1 year - Directors' loans",
    42: "Creditors less than 1 year - Directors' loans",
    43: "Debtors - Other debtors",  # VAT reclaim
    44: "Creditors less than 1 year - Other taxes and social security",
    45: "Creditors less than 1 year - Accruals",
    46: "Creditors less than 1 year - Other creditors",
    47: "Creditors less than 1 year - Bank loans",
    48: "Creditors less than 1 year - Finance lease and HP contracts",
    49: "Creditors greater than 1 year - Finance lease and HP contracts",
    50: "Deferred tax - Brought forward",
    51: "Share capital - Brought forward",
    52: "Profit and loss account - Brought forward",
    53: "Turnover - Sales",
    54: "Cost of sales - Purchases",
    55: "Cost of sales - Carriage",
    56: "Administrative expenses - General - Insurance",
    57: "Administrative expenses - Employee costs - Directors' salaries",
    58: "Administrative expenses - Employee costs - Directors' salaries",
    59: "Administrative expenses - Employee costs - Wages and salaries",
    60: "Administrative expenses - Employee costs - Employer's NI",
    61: "Administrative expenses - General - Telephone and fax",
    62: "Administrative expenses - General - Stationery and printing",
    63: "Administrative expenses - Employee costs - Travel and subsistence",
    64: "Administrative expenses - Employee costs - Motor expenses",
    65: "Administrative expenses - General - Repairs and maintenance",
    66: "Administrative expenses - General - Software",
    67: "Administrative expenses - General - Sundry expenses",
    68: "Administrative expenses - Legal & professional - Accountancy fees",
    69: "Cost of sales - Commissions payable",
    70: "Administrative expenses - Employee costs - Entertaining",
    71: "Administrative expenses - General - Bank charges",
    72: "Administrative expenses - General - Depreciation",
    73: "Administrative expenses - General - Depreciation",
    74: "Administrative expenses - General - Depreciation",
    75: "Administrative expenses - General - Depreciation",
    76: "Administrative expenses - General - Depreciation",
    77: "Administrative expenses - General - Depreciation",
    78: "Interest payable - Other loans",
    79: "Interest payable - Finance leases and HP",
    80: "Gains and losses - Gain/loss on sale of tangible assets",
    81: "Gains and losses - Gain/loss on sale of tangible assets",
    82: "Profit and loss account - Equity dividends",
}


def money(v) -> float:
    try:
        return round(float(v or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def chart_names() -> set[str]:
    wb = load_workbook(CHART, data_only=True)
    ws = wb.active
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            names.add(str(row[0]).strip())
    return names


def parse_source():
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    out = []
    for r in range(11, 83):
        name = ws[f"C{r}"].value
        if not name or not str(name).strip():
            continue
        out.append(
            {
                "row": r,
                "source": str(name).strip(),
                "dr": money(ws[f"I{r}"].value),
                "cr": money(ws[f"K{r}"].value),
            }
        )
    return out


def main() -> int:
    names = chart_names()
    rows = parse_source()
    buckets: dict[str, float] = defaultdict(float)
    mapping_rows = []
    unknown = []
    for row in rows:
        iris = BY_ROW.get(row["row"])
        if not iris:
            unknown.append(row)
            continue
        if iris not in names:
            unknown.append({**row, "iris": iris})
            continue
        amt = round(row["dr"] - row["cr"], 2)
        buckets[iris] += amt
        mapping_rows.append(
            {
                "row": row["row"],
                "source": row["source"],
                "iris": iris,
                "amount": amt,
            }
        )
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Account", "Description", "Year End 31/03/2026"])
        for account, amt in sorted(buckets.items()):
            if abs(amt) < 0.005:
                continue
            w.writerow([account, account, f"{amt:.2f}"])
    with MAP_OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["row", "source", "iris", "amount"])
        w.writeheader()
        w.writerows(mapping_rows)
    net = round(sum(buckets.values()), 2)
    print(f"wrote {OUT}")
    print(f"mapped {len(mapping_rows)} lines -> {len(buckets)} IRIS accounts  net {net}")
    if unknown:
        print("UNMAPPED", unknown)
    return 1 if unknown or abs(net) > 0.02 else 0


if __name__ == "__main__":
    raise SystemExit(main())
