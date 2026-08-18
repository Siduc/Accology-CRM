"""Frans Distribution Ltd — YE 31/03/2026 draft pack, IRIS TB, sales invoice PDFs.

Source: VAT review pack prepared 7 August 2026
  Accology Limited/Working Papers/
  Ftans Distribution Limited - long pe 30 June 2026 - VAT return workings.xlsx

Bank-ledger proxy. Sales invoices did not exist; PDFs are raised from
receipts from Sixty Six Interiors Ltd and Sixty Six South Limited.
Does not file at IRIS / CH / HMRC.
"""

from __future__ import annotations

import csv
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
    PageBreak,
    KeepTogether,
)
from pypdf import PdfWriter

CHART = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents"
    r"\Practice\Working Papers\Accology Chart.xlsx"
)
VAT_SRC = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Accology Limited\Working Papers"
    r"\Ftans Distribution Limited - long pe 30 June 2026 - VAT return workings.xlsx"
)
CLIENT_DIR = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Frans Distribution Ltd"
)

INCORP = date(2025, 3, 3)
YE = date(2026, 3, 31)
NAVY = "1B365D"
GOLD = "C4A35A"
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CUSTOMERS = {
    "interiors": {
        "name": "Sixty Six Interiors Ltd",
        "number": "16072864",
        "vat": "505 4218 22",
        "addr": [
            "Riverside",
            "Mountbatten Way",
            "Congleton",
            "CW12 1DY",
        ],
        "match": "sixty six interiors",
    },
    "south": {
        "name": "Sixty Six South Limited",
        "number": "16187631",
        "vat": "486 9668 09",
        "addr": [
            "c/o Accology Limited",
            "Bolton Arena, Arena Approach",
            "Horwich, Bolton",
            "BL6 6LB",
        ],
        "match": "sixty six south",
    },
}


def money(v) -> float:
    try:
        return round(float(v or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def chart_names() -> set[str]:
    wb = load_workbook(CHART, data_only=True)
    ws = wb.active
    names = {str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
    wb.close()
    return names


def load_ledger() -> list[dict]:
    wb = load_workbook(VAT_SRC, data_only=True)
    ws = wb["Transaction Ledger"]
    rows = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=16, values_only=True):
        d = as_date(r[0])
        if not d:
            continue
        desc = str(r[4] or "")
        payee = str(r[14] or "")
        rows.append(
            {
                "date": d,
                "amount": money(r[1]),
                "balance": money(r[2]),
                "desc": desc,
                "direction": str(r[5] or ""),
                "treat": str(r[6] or ""),
                "net_sales": money(r[8]),
                "output_vat": money(r[9]),
                "box7": money(r[10]),
                "input_vat": money(r[11]),
                "payee": payee,
                "activity": str(r[15] or ""),
            }
        )
    wb.close()
    return rows


def customer_key(row: dict) -> str | None:
    blob = f"{row['payee']} {row['desc']}".lower()
    if "sixty six interiors" in blob:
        return "interiors"
    if "sixty six south" in blob:
        return "south"
    return None


def service_line(row: dict) -> str:
    desc = row["desc"]
    # Bank memo after last comma is the short service label
    parts = [p.strip() for p in desc.split(",") if p.strip()]
    label = parts[-1] if parts else "Deliveries"
    mapping = {
        "Deliveries": "Delivery / haulage services",
        "Haulage": "Haulage services",
        "Sofa Clearance": "Sofa clearance / delivery",
        "Panels": "Panel delivery",
        "Ref Panel": "Refused / return panel delivery",
        "Inv": "Delivery / haulage services",
        "Stock": "Stock delivery",
        "Clearance": "Clearance / delivery",
    }
    return mapping.get(label, label)


def classify_purchase(row: dict) -> str:
    act = (row["activity"] or "").lower()
    desc = (row["desc"] or "").lower()
    payee = (row["payee"] or "").lower()
    if "ricky" in payee:
        if "re pay" in desc or "repay" in desc:
            return "dla"
        if "fuel" in desc or act == "fuel":
            return "Administrative expenses - Employee costs - Motor expenses"
        if "wages" in desc or act == "wages / payroll":
            return "Administrative expenses - Employee costs - Directors' salaries"
        if "station" in desc:
            return "Administrative expenses - General - Stationery and printing"
        if "expense" in desc:
            return "Administrative expenses - General - Sundry expenses"
        return "dla"
    if act == "deliveries":
        return "Cost of sales - Carriage"
    if act == "stock / inventory":
        return "Cost of sales - Purchases"
    if act == "wages / payroll":
        return "Administrative expenses - Employee costs - Wages and salaries"
    if act == "vehicles / rent":
        return "Administrative expenses - Premises costs - Rent"
    if act == "software / online services":
        return "Administrative expenses - General - Software"
    if act == "advertising":
        return "Administrative expenses - Legal & professional - Advertising and PR"
    if act == "communications":
        return "Administrative expenses - General - Telephone and fax"
    if act == "fuel":
        return "Administrative expenses - Employee costs - Motor expenses"
    if act == "waste services":
        return "Administrative expenses - General - Sundry expenses"
    if act == "professional services":
        return "Administrative expenses - Legal & professional - Other legal and professional"
    if act == "labels / packaging":
        return "Cost of sales - Other direct costs"
    if act == "bank adjustments":
        return "suspense"
    if act == "repayments":
        return "dla"
    if act == "director payments":
        return "dla"
    if "mint formations" in payee:
        return "Administrative expenses - Legal & professional - Other legal and professional"
    if act in ("general purchases", "card purchases / review", "expenses", "administration", "other / review", "direct debits / review"):
        return "Administrative expenses - General - Sundry expenses"
    return "Administrative expenses - General - Sundry expenses"


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(color="FFFFFF", bold=True, name="Calibri")
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font


def write_money(ws, r, c, value):
    cell = ws.cell(r, c, round(float(value or 0), 2))
    cell.number_format = '#,##0.00;(#,##0.00);"—"'
    cell.alignment = Alignment(horizontal="right")


def write_iris_csv(path: Path, buckets: dict[str, float], names: set[str]) -> tuple[float, list[str]]:
    unknown = [a for a in buckets if a not in names and abs(buckets[a]) >= 0.005]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Account", "Description", "Year End 31/03/2026"])
        for account in sorted(buckets):
            amt = round(buckets[account], 2)
            if abs(amt) < 0.005:
                continue
            w.writerow([account, account, f"{amt:.2f}"])
    return round(sum(buckets.values()), 2), unknown


def build_invoices(all_rows: list[dict]) -> list[dict]:
    sales = []
    for row in all_rows:
        if row["direction"] != "Credit":
            continue
        key = customer_key(row)
        if not key:
            continue
        sales.append((row["date"], key, row))
    sales.sort(key=lambda x: (x[0], x[1], x[2]["amount"]))
    invoices = []
    for i, (d, key, row) in enumerate(sales, start=1):
        invoices.append(
            {
                "number": f"FD-{i:04d}",
                "date": d,
                "customer": key,
                "gross": row["amount"],
                "net": row["net_sales"],
                "vat": row["output_vat"],
                "service": service_line(row),
                "bank_ref": row["desc"],
                "in_year": d <= YE,
            }
        )
    return invoices


def draw_invoice_pdf(path: Path, inv: dict) -> None:
    cust = CUSTOMERS[inv["customer"]]
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "InvTitle",
        parent=styles["Heading1"],
        fontName="Times-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1B365D"),
        spaceAfter=2,
    )
    sub = ParagraphStyle(
        "InvSub",
        parent=styles["Normal"],
        fontName="Times-Roman",
        fontSize=9,
        textColor=colors.HexColor("#444444"),
        leading=12,
    )
    body = ParagraphStyle(
        "InvBody",
        parent=styles["Normal"],
        fontName="Times-Roman",
        fontSize=10,
        leading=13,
    )
    bold = ParagraphStyle(
        "InvBold",
        parent=body,
        fontName="Times-Bold",
    )
    small = ParagraphStyle(
        "InvSmall",
        parent=body,
        fontSize=8,
        textColor=colors.HexColor("#555555"),
        leading=10,
    )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=f"{inv['number']} {cust['name']}",
        author="Frans Distribution Ltd",
    )
    story = []
    story.append(Paragraph("FRANS DISTRIBUTION LTD", title))
    story.append(
        Paragraph(
            "Registered office: 2nd Floor College House, 17 King Edwards Road, Ruislip HA4 7AE<br/>"
            "Trading: 20 Woodfin Croft, Chelford, Macclesfield SK11 9SN<br/>"
            "Company number 16286556 &nbsp;&nbsp; VAT 522 0387 23",
            sub,
        )
    )
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#1B365D")))
    story.append(Spacer(1, 8))
    story.append(Paragraph("TAX INVOICE", ParagraphStyle("H2", parent=title, fontSize=14)))
    meta = [
        [
            Paragraph("<b>Invoice to</b>", body),
            Paragraph("<b>Invoice</b>", body),
        ],
        [
            Paragraph(
                f"{cust['name']}<br/>"
                + "<br/>".join(cust["addr"])
                + f"<br/>Company {cust['number']}<br/>VAT {cust['vat']}",
                body,
            ),
            Paragraph(
                f"Invoice number: <b>{inv['number']}</b><br/>"
                f"Invoice / tax point: <b>{inv['date'].strftime('%d %B %Y')}</b><br/>"
                f"Payment: Paid on {inv['date'].strftime('%d %B %Y')}<br/>"
                "VAT rate: 20% (amount received is VAT-inclusive)",
                body,
            ),
        ],
    ]
    t = Table(meta, colWidths=[90 * mm, 80 * mm])
    t.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 12))
    hdr = [
        Paragraph("<b>Description</b>", body),
        Paragraph("<b>Net £</b>", body),
        Paragraph("<b>VAT £</b>", body),
        Paragraph("<b>Gross £</b>", body),
    ]
    line = [
        Paragraph(f"{inv['service']}<br/><font size='8'>Bank reference: {inv['bank_ref']}</font>", body),
        Paragraph(f"{inv['net']:,.2f}", body),
        Paragraph(f"{inv['vat']:,.2f}", body),
        Paragraph(f"{inv['gross']:,.2f}", body),
    ]
    tot = [
        Paragraph("<b>Total — paid in full</b>", bold),
        Paragraph(f"<b>{inv['net']:,.2f}</b>", bold),
        Paragraph(f"<b>{inv['vat']:,.2f}</b>", bold),
        Paragraph(f"<b>{inv['gross']:,.2f}</b>", bold),
    ]
    grid = Table([hdr, line, tot], colWidths=[100 * mm, 23 * mm, 23 * mm, 24 * mm])
    grid.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B365D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#F4F1EA")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(grid)
    story.append(Spacer(1, 14))
    story.append(
        Paragraph(
            "This invoice is raised from the cleared bank receipt. "
            "The tax point used is the date the payment reached Frans Distribution Ltd. "
            "No balance remains outstanding.",
            body,
        )
    )
    story.append(Spacer(1, 10))
    story.append(
        Paragraph(
            "Bank payment received in full. Please retain this invoice for your VAT records.",
            bold,
        )
    )
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#AAAAAA")))
    story.append(
        Paragraph(
            "Frans Distribution Ltd is a company registered in England and Wales. "
            "VAT invoices issued under VAT registration 522 0387 23.",
            small,
        )
    )
    doc.build(story)


def write_combined(pdfs: list[Path], dest: Path) -> None:
    w = PdfWriter()
    for p in pdfs:
        w.append(str(p))
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("wb") as f:
        w.write(f)


def write_pack(d: dict, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Frans Distribution Ltd"
    cover["A1"].font = Font(name="Calibri", bold=True, size=18, color=NAVY)
    cover["A2"] = "Draft accounts working pack — first period ended 31 March 2026"
    cover["A3"] = (
        "Prepared from the 7 August 2026 VAT review pack "
        "(bank ledger 13 October 2025 to 30 June 2026), cut at 31 March 2026."
    )
    cover["A4"] = (
        "NOT an IRIS statutory set. Import the IRIS Elements CSV after queries. "
        "Do not file at Companies House / HMRC without client approval."
    )
    cover["A4"].font = Font(italic=True, color="833C0C")
    r = 6
    cover.cell(r, 1, "Draft figures")
    cover.cell(r, 1).font = Font(bold=True, color=NAVY)
    r += 1
    for label, val in d["cover_stats"]:
        cover.cell(r, 1, label)
        if isinstance(val, (int, float)):
            write_money(cover, r, 2, val)
        else:
            cover.cell(r, 2, val)
        r += 1
    r += 1
    cover.cell(r, 1, "Open queries")
    cover.cell(r, 1).font = Font(bold=True, color=NAVY)
    r += 1
    for i, q in enumerate(d["queries"], 1):
        cover.cell(r, 1, f"{i}. {q}")
        cover.cell(r, 1).alignment = Alignment(wrap_text=True)
        cover.row_dimensions[r].height = 30
        r += 1
    cover.column_dimensions["A"].width = 112
    cover.column_dimensions["B"].width = 18

    qws = wb.create_sheet("Queries")
    qws["A1"] = "No"
    qws["B1"] = "Query"
    style_header(qws, 1, 2)
    for i, q in enumerate(d["queries"], 1):
        qws.cell(i + 1, 1, i)
        qws.cell(i + 1, 2, q)
        qws.cell(i + 1, 2).alignment = Alignment(wrap_text=True, vertical="top")
        qws.row_dimensions[i + 1].height = 34
    qws.column_dimensions["A"].width = 6
    qws.column_dimensions["B"].width = 112

    led = wb.create_sheet("YE bank ledger")
    led["A1"] = "Date"
    led["B1"] = "Direction"
    led["C1"] = "Payee / description"
    led["D1"] = "Activity"
    led["E1"] = "Gross"
    led["F1"] = "Accology / IRIS"
    style_header(led, 1, 6)
    for i, row in enumerate(d["ye_rows_out"], 2):
        led.cell(i, 1, row["date"].isoformat())
        led.cell(i, 2, row["direction"])
        led.cell(i, 3, row["desc"])
        led.cell(i, 4, row["activity"])
        write_money(led, i, 5, row["amount"])
        led.cell(i, 6, row["iris"])
    for col, w in (("A", 14), ("B", 12), ("C", 62), ("D", 28), ("E", 14), ("F", 56)):
        led.column_dimensions[col].width = w

    invs = wb.create_sheet("Sales invoices")
    invs["A1"] = "Invoice"
    invs["B1"] = "Date"
    invs["C1"] = "Customer"
    invs["D1"] = "Service"
    invs["E1"] = "Net"
    invs["F1"] = "VAT"
    invs["G1"] = "Gross"
    invs["H1"] = "In YE 31/3/2026?"
    style_header(invs, 1, 8)
    for i, inv in enumerate(d["invoices"], 2):
        invs.cell(i, 1, inv["number"])
        invs.cell(i, 2, inv["date"].isoformat())
        invs.cell(i, 3, CUSTOMERS[inv["customer"]]["name"])
        invs.cell(i, 4, inv["service"])
        write_money(invs, i, 5, inv["net"])
        write_money(invs, i, 6, inv["vat"])
        write_money(invs, i, 7, inv["gross"])
        invs.cell(i, 8, "Yes" if inv["in_year"] else "After year end")
    for col, w in (("A", 12), ("B", 14), ("C", 32), ("D", 32), ("E", 14), ("F", 12), ("G", 14), ("H", 20)):
        invs.column_dimensions[col].width = w

    pl = wb.create_sheet("Draft P&L")
    pl["A1"] = "Profit and loss — 3 March 2025 to 31 March 2026"
    pl["A1"].font = Font(bold=True, size=14, color=NAVY)
    pl["A3"] = "Line"
    pl["B3"] = "£"
    style_header(pl, 3, 2)
    for i, (label, val) in enumerate(d["pl_lines"], 4):
        pl.cell(i, 1, label)
        write_money(pl, i, 2, val)
        if label in ("Turnover", "Gross profit", "Operating profit", "Profit before tax", "Profit for the year (before tax)"):
            pl.cell(i, 1).font = Font(bold=True)
    pl.column_dimensions["A"].width = 62
    pl.column_dimensions["B"].width = 16

    bs = wb.create_sheet("Draft balance sheet")
    bs["A1"] = "Balance sheet — 31 March 2026"
    bs["A1"].font = Font(bold=True, size=14, color=NAVY)
    r = 3
    for title, rows in d["bs_sections"]:
        bs.cell(r, 1, title)
        bs.cell(r, 2, "£")
        style_header(bs, r, 2)
        r += 1
        for label, val in rows:
            bs.cell(r, 1, label)
            write_money(bs, r, 2, val)
            if label.lower().startswith("total") or label.lower().startswith("net") or "funds" in label.lower():
                bs.cell(r, 1).font = Font(bold=True)
            r += 1
        r += 1
    bs.column_dimensions["A"].width = 62
    bs.column_dimensions["B"].width = 16

    iris = wb.create_sheet("IRIS Elements TB")
    iris["A1"] = "Account"
    iris["B1"] = "Description"
    iris["C1"] = "Year ended 31 March 2026"
    style_header(iris, 1, 3)
    net = 0.0
    r = 2
    for account in sorted(d["buckets"]):
        amt = round(d["buckets"][account], 2)
        if abs(amt) < 0.005:
            continue
        iris.cell(r, 1, account)
        iris.cell(r, 2, account)
        write_money(iris, r, 3, amt)
        net += amt
        r += 1
    iris.cell(r, 1, "Net (must be 0.00)")
    iris.cell(r, 1).font = Font(bold=True)
    write_money(iris, r, 3, net)
    iris.column_dimensions["A"].width = 62
    iris.column_dimensions["B"].width = 62
    iris.column_dimensions["C"].width = 22
    wb.save(dest)


def write_queries_md(path: Path, queries: list[str]) -> None:
    lines = [
        "# Frans Distribution Ltd — draft queries",
        "",
        "First period ended 31 March 2026. Draft only. Confirm before IRIS finals / CH / HMRC.",
        "",
    ]
    for i, q in enumerate(queries, 1):
        lines.append(f"{i}. {q}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    names = chart_names()
    rows = load_ledger()
    invoices = build_invoices(rows)
    ye_rows = [r for r in rows if r["date"] <= YE]

    buckets: dict[str, float] = defaultdict(float)
    ye_out = []
    sales_net = {"interiors": 0.0, "south": 0.0}
    sales_vat = {"interiors": 0.0, "south": 0.0}
    sales_gross = {"interiors": 0.0, "south": 0.0}

    # Sales invoices dated in the year
    for inv in invoices:
        if not inv["in_year"]:
            continue
        buckets["Turnover - Sales"] -= inv["net"]
        buckets["Creditors less than 1 year - Other taxes and social security"] -= inv["vat"]
        buckets["Cash - Cash at bank and in hand"] += inv["gross"]
        sales_net[inv["customer"]] += inv["net"]
        sales_vat[inv["customer"]] += inv["vat"]
        sales_gross[inv["customer"]] += inv["gross"]

    dla = 0.0
    suspense = 0.0
    other_income = 0.0

    for row in ye_rows:
        key = customer_key(row)
        if row["direction"] == "Credit" and key:
            ye_out.append({**row, "iris": "Turnover - Sales / VAT / Bank"})
            continue
        if row["direction"] == "Credit":
            desc = (row["desc"] + " " + row["payee"]).lower()
            if "adjustment" in desc:
                # Opening / director introduction, not a sale
                buckets["Cash - Cash at bank and in hand"] += row["amount"]
                buckets["Creditors less than 1 year - Directors' loans"] -= row["amount"]
                dla -= row["amount"]
                ye_out.append({**row, "iris": "Directors' loans (capital introduced)"})
                continue
            if "cashback" in desc:
                buckets["Cash - Cash at bank and in hand"] += row["amount"]
                buckets["Other operating income"] -= row["amount"]
                other_income += row["amount"]
                ye_out.append({**row, "iris": "Other operating income"})
                continue
            if "failed payment" in desc:
                buckets["Cash - Cash at bank and in hand"] += row["amount"]
                buckets["Cost of sales - Carriage"] -= row["amount"]
                ye_out.append({**row, "iris": "Cost of sales - Carriage (failed payment reverse)"})
                continue
            # Residual credits (e.g. Broker Experts refund)
            buckets["Cash - Cash at bank and in hand"] += row["amount"]
            net = row["net_sales"] or round(row["amount"] / 1.2, 2)
            vat = row["output_vat"] or round(row["amount"] - net, 2)
            buckets["Other operating income"] -= net
            buckets["Creditors less than 1 year - Other taxes and social security"] -= vat
            other_income += net
            ye_out.append({**row, "iris": "Other operating income / VAT"})
            continue

        # Debits
        iris = classify_purchase(row)
        paid = abs(row["amount"])
        buckets["Cash - Cash at bank and in hand"] -= paid
        if iris == "dla":
            buckets["Creditors less than 1 year - Directors' loans"] += paid
            dla += paid
            ye_out.append({**row, "iris": "Directors' loans"})
            continue
        if iris == "suspense":
            buckets["Debtors - Other debtors"] += paid
            suspense += paid
            ye_out.append({**row, "iris": "Debtors - Other debtors (bank adjustments — query)"})
            continue
        vat = row["input_vat"]
        net = row["box7"] if abs(row["box7"]) > 0.004 else round(paid - vat, 2)
        if vat > 0.004:
            buckets["Creditors less than 1 year - Other taxes and social security"] += vat
        buckets[iris] += net
        ye_out.append({**row, "iris": iris})

    # Year-end accountancy accrual — CRM Accounts fee £2,500
    buckets["Administrative expenses - Legal & professional - Accountancy fees"] += 2500.00
    buckets["Creditors less than 1 year - Accruals"] -= 2500.00

    # Round
    for k in list(buckets):
        buckets[k] = round(buckets[k], 2)

    turnover = -buckets.get("Turnover - Sales", 0.0)
    carriage = buckets.get("Cost of sales - Carriage", 0.0)
    purchases = buckets.get("Cost of sales - Purchases", 0.0)
    other_cos = buckets.get("Cost of sales - Other direct costs", 0.0)
    cos = round(carriage + purchases + other_cos, 2)
    gross = round(turnover - cos, 2)
    dirsals = buckets.get("Administrative expenses - Employee costs - Directors' salaries", 0.0)
    wages = buckets.get("Administrative expenses - Employee costs - Wages and salaries", 0.0)
    motor = buckets.get("Administrative expenses - Employee costs - Motor expenses", 0.0)
    rent = buckets.get("Administrative expenses - Premises costs - Rent", 0.0)
    software = buckets.get("Administrative expenses - General - Software", 0.0)
    ads = buckets.get("Administrative expenses - Legal & professional - Advertising and PR", 0.0)
    phone = buckets.get("Administrative expenses - General - Telephone and fax", 0.0)
    staty = buckets.get("Administrative expenses - General - Stationery and printing", 0.0)
    sundry = buckets.get("Administrative expenses - General - Sundry expenses", 0.0)
    legal = buckets.get("Administrative expenses - Legal & professional - Other legal and professional", 0.0)
    acc = buckets.get("Administrative expenses - Legal & professional - Accountancy fees", 0.0)
    oth_inc = -buckets.get("Other operating income", 0.0)
    overheads = round(dirsals + wages + motor + rent + software + ads + phone + staty + sundry + legal + acc, 2)
    profit = round(gross - overheads + oth_inc, 2)

    bank = buckets.get("Cash - Cash at bank and in hand", 0.0)
    vat_cr = -buckets.get("Creditors less than 1 year - Other taxes and social security", 0.0)
    accruals = -buckets.get("Creditors less than 1 year - Accruals", 0.0)
    dla_cr = -buckets.get("Creditors less than 1 year - Directors' loans", 0.0)
    oth_dr = buckets.get("Debtors - Other debtors", 0.0)
    net_assets = round(bank + oth_dr - vat_cr - accruals - dla_cr, 2)
    funds = profit  # no opening reserves / share capital on books

    queries = [
        "Accounts are a bank-ledger proxy from the 7 August 2026 VAT pack (filename typo: Ftans). Tax points are bank dates, not invoice dates. Sales invoices have now been raised from those receipts — confirm Ricky / Alan accept them.",
        "Incorporated 3 March 2025 (16286556). First bank movement 13 October 2025. Confirm there was no earlier activity and that the first period is 3 March 2025 to 31 March 2026.",
        "Registered office on Companies House is College House, Ruislip (Mint Formations). CRM trading address is 20 Woodfin Croft, Chelford SK11 9SN. Confirm which goes on the statutory accounts.",
        "Opening Adjustment CREDIT £22,085.00 on 13 October 2025 treated as director capital introduced (DLA credit), not a sale. The VAT pack had treated it as standard-rated sales. Confirm with Ricky.",
        "Bank Adjustment DEBITS £16,655.00 parked in other debtors / suspense. Need an explanation before finals.",
        "VAT on the draft is computed from receipts and the VAT-pack input ticks, not from supplier invoices. Input VAT claimed in the long-period pack was £136,979 to 30 June; year-end slice is on this pack. Do not file the VAT return or CT without invoices.",
        "Delivery costs are mostly FPS/MOB to named drivers and van-and-man firms. Confirm CIS / self-employed status and that valid VAT invoices exist before reclaiming input VAT.",
        "Ricky Stray 'wages' payments have no RTI / payroll file on the papers. Draft treats labelled wages as directors' remuneration and repayments / unexplained Ricky items as DLA.",
        "Sophia Fidler £420 and Grace Hall £408.50 look like staff wages. Need payroll records.",
        "Empire £47,320 (stock / other) — confirm what was bought and whether any stock remains at 31 March 2026. No stock is on the draft balance sheet.",
        "L8 Rent & Hire and Mrs Nicola Carter £3,000 — vehicle / van hire. Need contracts and VAT invoices.",
        "Juniper Wilmslow (31 card spends), Mottram Leisure DD, Peking Garden, MyGP Clinic, Booking.com hotel — confirm wholly business.",
        "Share capital is not on the bank ledger. Draft has no called-up share capital line. Confirm allotment (£1?).",
        "Accountancy accrual £2,500 is the CRM Accounts job fee. Confirm engagement.",
        "Corporation tax not computed (need capital allowances, director remuneration, and whether this is a micro-entity).",
        "The VAT pack continues to 30 June 2026. April–June receipts from Sixty Six Interiors / South are on the invoice PDFs but are NOT in these accounts.",
        "Failed payment credit £400 treated as a reverse of delivery costs. Broker Experts credit £333.23 treated as other income plus VAT — confirm.",
    ]

    cover_stats = [
        ("Incorporated", "3 March 2025"),
        ("First bank activity", "13 October 2025"),
        ("Year end", "31 March 2026"),
        ("Turnover (Sixty Six Interiors + South)", turnover),
        ("Gross profit", gross),
        ("Draft profit before tax", profit),
        ("Cash at bank 31 March 2026", bank),
        ("VAT creditor (draft)", vat_cr),
        ("Directors' loan (credit = company owes Ricky)", dla_cr),
        ("Net assets / shareholders' funds", net_assets),
        ("IRIS TB net (must be 0.00)", 0.0),
        ("Sales invoices raised (all dates)", len(invoices)),
        ("of which in this year", sum(1 for i in invoices if i["in_year"])),
    ]

    pl_lines = [
        ("Turnover — Sixty Six Interiors Ltd", sales_net["interiors"]),
        ("Turnover — Sixty Six South Limited", sales_net["south"]),
        ("Turnover", turnover),
        ("Delivery / subcontract haulage", -carriage),
        ("Purchases / stock", -purchases),
        ("Labels / packaging", -other_cos),
        ("Cost of sales", -cos),
        ("Gross profit", gross),
        ("Directors' remuneration (Ricky Stray)", -dirsals),
        ("Wages and salaries", -wages),
        ("Motor / fuel", -motor),
        ("Van / vehicle hire", -rent),
        ("Software and online services", -software),
        ("Advertising", -ads),
        ("Telephone and communications", -phone),
        ("Stationery", -staty),
        ("Sundry / card / other review", -sundry),
        ("Legal and professional", -legal),
        ("Accountancy accrual (draft)", -acc),
        ("Other operating income", oth_inc),
        ("Profit before tax", profit),
        ("Tax (not computed — see Queries)", 0.0),
        ("Profit for the year (before tax)", profit),
    ]

    bs_sections = [
        (
            "Current assets",
            [
                ("Cash at bank", bank),
                ("Other debtors (bank adjustments — query)", oth_dr),
                ("Total current assets", round(bank + oth_dr, 2)),
            ],
        ),
        (
            "Creditors: amounts falling due within one year",
            [
                ("VAT (output less input, draft)", vat_cr),
                ("Accruals (accountancy)", accruals),
                ("Directors' loan — Ricky Stray", dla_cr),
                ("Total creditors", round(vat_cr + accruals + dla_cr, 2)),
            ],
        ),
        (
            "Capital and reserves",
            [
                ("Profit for the period (before tax)", profit),
                ("Shareholders' funds", funds),
                ("Net assets (draft)", net_assets),
            ],
        ),
    ]

    # Copy VAT source into the client pack
    src_dir = CLIENT_DIR / "Current" / "Source"
    src_dir.mkdir(parents=True, exist_ok=True)
    copied = src_dir / "Frans Distribution Limited - long pe 30 June 2026 - VAT return workings.xlsx"
    if VAT_SRC.exists():
        shutil.copy2(VAT_SRC, copied)

    pack = CLIENT_DIR / "Current" / "Working Papers" / "Frans Distribution Limited - 31 March 2026 - Draft accounts pack.xlsx"
    iris = CLIENT_DIR / "Current" / "IRIS Import" / "2026-03-31 IRIS Elements TB.csv"
    qmd = CLIENT_DIR / "Current" / "Working Papers" / "Frans Distribution Limited - 31 March 2026 - Queries.md"
    inv_root = CLIENT_DIR / "Current" / "Invoices"
    inv_root.mkdir(parents=True, exist_ok=True)

    net, unknown = write_iris_csv(iris, buckets, names)
    write_queries_md(qmd, queries)
    write_pack(
        {
            "queries": queries,
            "cover_stats": cover_stats[:-3] + [("IRIS TB net (must be 0.00)", net), cover_stats[-2], cover_stats[-1]],
            "ye_rows_out": ye_out,
            "invoices": invoices,
            "pl_lines": pl_lines,
            "bs_sections": bs_sections,
            "buckets": buckets,
        },
        pack,
    )

    # Invoice PDFs
    pdfs_by = {"interiors": [], "south": []}
    for inv in invoices:
        folder = inv_root / CUSTOMERS[inv["customer"]]["name"]
        fname = f"{inv['number']} {CUSTOMERS[inv['customer']]['name']} {inv['date'].isoformat()}.pdf"
        dest = folder / fname
        draw_invoice_pdf(dest, inv)
        pdfs_by[inv["customer"]].append(dest)

    for key, pdfs in pdfs_by.items():
        write_combined(
            pdfs,
            inv_root / f"Frans Distribution Ltd - all invoices to {CUSTOMERS[key]['name']}.pdf",
        )

    # Register CSV
    reg = inv_root / "Frans Distribution Ltd - sales invoice register.csv"
    with reg.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Invoice", "Date", "Customer", "Service", "Net", "VAT", "Gross", "In YE 31/3/2026", "Bank reference"])
        for inv in invoices:
            w.writerow(
                [
                    inv["number"],
                    inv["date"].isoformat(),
                    CUSTOMERS[inv["customer"]]["name"],
                    inv["service"],
                    f"{inv['net']:.2f}",
                    f"{inv['vat']:.2f}",
                    f"{inv['gross']:.2f}",
                    "Yes" if inv["in_year"] else "No",
                    inv["bank_ref"],
                ]
            )

    print(f"pack={pack}")
    print(f"iris={iris} net={net} unknown={unknown}")
    print(f"profit={profit} bank={bank} vat={vat_cr} dla={dla_cr} net_assets={net_assets} funds={funds}")
    print(f"invoices={len(invoices)} interiors={len(pdfs_by['interiors'])} south={len(pdfs_by['south'])}")
    print(f"queries={qmd}")
    print(f"register={reg}")
    if unknown:
        print("UNMAPPED", unknown)
        return 1
    if abs(net) > 0.05:
        print("TB OUT OF BALANCE")
        return 1
    if abs(net_assets - funds) > 0.05:
        print(f"BS/equity mismatch na={net_assets} funds={funds}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
