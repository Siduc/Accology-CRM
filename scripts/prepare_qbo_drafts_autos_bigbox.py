"""Finish Production Autos YE 28/02/2026 and Big Box Art YE 31/12/2025 to draft.

Reads the current QuickBooks working-paper packs, recodes to the Accology
Chart (IRIS Elements / Taxfiler names), posts year-end journals that can be
calculated, and writes:

  Current/Working Papers/* Draft accounts pack.xlsx
  Current/IRIS Import/YYYY-MM-DD IRIS Elements TB.csv
  Current/IRIS Import/YYYY-MM-DD Accology chart mapping.csv
  Current/Working Papers/* Queries.md

Does not submit to IRIS or Companies House. Journals that need a client
answer stay on the query list instead of being invented.
"""

from __future__ import annotations

import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

CHART = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents"
    r"\Practice\Working Papers\Accology Chart.xlsx"
)
CLIENTS = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
)

NAVY = "052891"
GOLD = "F4B809"
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def money(v) -> float:
    try:
        return round(float(v or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def chart_names() -> set[str]:
    wb = load_workbook(CHART, data_only=True)
    ws = wb.active
    names = {str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
    wb.close()
    return names


def style_header(ws, row: int, cols: int, title: str | None = None) -> None:
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(color="FFFFFF", bold=True, name="Calibri")
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
    if title:
        ws.cell(row, 1, title)


def write_money(ws, r: int, c: int, value) -> None:
    cell = ws.cell(r, c, round(float(value or 0), 2))
    cell.number_format = '#,##0.00;(#,##0.00);"—"'
    cell.alignment = Alignment(horizontal="right")


def apply_jnls(buckets: dict[str, float], journals: list[dict]) -> None:
    for j in journals:
        buckets[j["account"]] = round(buckets.get(j["account"], 0.0) + j["amount"], 2)


def write_iris_csv(path: Path, ye_header: str, buckets: dict[str, float], names: set[str]) -> tuple[float, list[str]]:
    unknown = [a for a in buckets if a not in names and abs(buckets[a]) >= 0.005]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Account", "Description", ye_header])
        for account in sorted(buckets):
            amt = round(buckets[account], 2)
            if abs(amt) < 0.005:
                continue
            w.writerow([account, account, f"{amt:.2f}"])
    net = round(sum(buckets.values()), 2)
    return net, unknown


def write_mapping_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["source", "iris", "qbo_amount", "journals", "final"]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def write_pack(
    dest: Path,
    *,
    company: str,
    subtitle: str,
    source_note: str,
    ye_label: str,
    qbo_rows: list[dict],
    journals: list[dict],
    buckets: dict[str, float],
    queries: list[str],
    cover_stats: list[tuple[str, object]],
    pl_lines: list[tuple[str, float]],
    bs_sections: list[tuple[str, list[tuple[str, float]]]],
) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = company
    cover["A1"].font = Font(name="Calibri", bold=True, size=18, color=NAVY)
    cover["A2"] = subtitle
    cover["A2"].font = Font(name="Calibri", size=12)
    cover["A3"] = source_note
    cover["A4"] = (
        "NOT an IRIS statutory set. Import the IRIS Elements CSV, review queries, "
        "get client approval, then file in IRIS."
    )
    cover["A4"].font = Font(name="Calibri", italic=True, color="833C0C")
    r = 6
    cover.cell(r, 1, "Draft figures")
    cover.cell(r, 1).font = Font(bold=True, color=NAVY)
    r += 1
    for label, val in cover_stats:
        cover.cell(r, 1, label)
        if isinstance(val, (int, float)):
            write_money(cover, r, 2, val)
        else:
            cover.cell(r, 2, val)
        r += 1
    r += 1
    cover.cell(r, 1, "Open queries")
    cover.cell(r, 1).font = Font(bold=True, color=NAVY)
    r += 1
    for i, q in enumerate(queries, 1):
        cover.cell(r, 1, f"{i}. {q}")
        cover.cell(r, 1).alignment = Alignment(wrap_text=True)
        cover.row_dimensions[r].height = 32
        r += 1
    cover.column_dimensions["A"].width = 110
    cover.column_dimensions["B"].width = 18

    qws = wb.create_sheet("Queries")
    qws["A1"] = "No"
    qws["B1"] = "Query"
    qws["C1"] = "Blocks IRIS final?"
    style_header(qws, 1, 3)
    for i, q in enumerate(queries, 1):
        qws.cell(i + 1, 1, i)
        qws.cell(i + 1, 2, q)
        qws.cell(i + 1, 2).alignment = Alignment(wrap_text=True, vertical="top")
        qws.cell(i + 1, 3, "Yes — confirm before filing")
        qws.row_dimensions[i + 1].height = 36
    qws.column_dimensions["A"].width = 6
    qws.column_dimensions["B"].width = 110
    qws.column_dimensions["C"].width = 28

    tb = wb.create_sheet("QBO TB")
    tb["A1"] = "QuickBooks account"
    tb["B1"] = "Debit"
    tb["C1"] = "Credit"
    tb["D1"] = "Accology / IRIS name"
    style_header(tb, 1, 4)
    r = 2
    tot_dr = tot_cr = 0.0
    for row in qbo_rows:
        tb.cell(r, 1, row["source"])
        if row["amount"] >= 0:
            write_money(tb, r, 2, row["amount"])
            tot_dr += row["amount"]
        else:
            write_money(tb, r, 3, -row["amount"])
            tot_cr += -row["amount"]
        tb.cell(r, 4, row["iris"])
        r += 1
    tb.cell(r, 1, "Total")
    tb.cell(r, 1).font = Font(bold=True)
    write_money(tb, r, 2, tot_dr)
    write_money(tb, r, 3, tot_cr)
    tb.column_dimensions["A"].width = 56
    tb.column_dimensions["B"].width = 16
    tb.column_dimensions["C"].width = 16
    tb.column_dimensions["D"].width = 62

    jws = wb.create_sheet("Draft journals")
    jws["A1"] = "Jnl"
    jws["B1"] = "Accology / IRIS account"
    jws["C1"] = "Debit"
    jws["D1"] = "Credit"
    jws["E1"] = "Narrative"
    style_header(jws, 1, 5)
    r = 2
    for j in journals:
        jws.cell(r, 1, j["jnl"])
        jws.cell(r, 2, j["account"])
        if j["amount"] >= 0:
            write_money(jws, r, 3, j["amount"])
        else:
            write_money(jws, r, 4, -j["amount"])
        jws.cell(r, 5, j["narr"])
        r += 1
    jws.column_dimensions["A"].width = 10
    jws.column_dimensions["B"].width = 62
    jws.column_dimensions["C"].width = 14
    jws.column_dimensions["D"].width = 14
    jws.column_dimensions["E"].width = 72

    pl = wb.create_sheet("Draft P&L")
    pl["A1"] = f"Profit and loss — {ye_label}"
    pl["A1"].font = Font(bold=True, size=14, color=NAVY)
    pl["A3"] = "Line"
    pl["B3"] = "£"
    style_header(pl, 3, 2)
    bold_labels = {
        "Gross profit",
        "Operating profit",
        "Profit before tax",
        "Profit for the year (before tax)",
        "Turnover",
    }
    for i, (label, val) in enumerate(pl_lines, start=4):
        pl.cell(i, 1, label)
        write_money(pl, i, 2, val)
        if label in bold_labels:
            pl.cell(i, 1).font = Font(bold=True)
    pl.column_dimensions["A"].width = 62
    pl.column_dimensions["B"].width = 16

    bs = wb.create_sheet("Draft balance sheet")
    bs["A1"] = f"Balance sheet — {ye_label}"
    bs["A1"].font = Font(bold=True, size=14, color=NAVY)
    r = 3
    for title, rows in bs_sections:
        bs.cell(r, 1, title)
        bs.cell(r, 2, "£")
        style_header(bs, r, 2)
        r += 1
        for label, val in rows:
            bs.cell(r, 1, label)
            write_money(bs, r, 2, val)
            if label.lower().startswith("total") or label.lower().startswith("net") or "funds" in label.lower():
                bs.cell(r, 1).font = Font(bold=True)
            r += 1
        r += 1
    bs.column_dimensions["A"].width = 62
    bs.column_dimensions["B"].width = 16

    iris = wb.create_sheet("IRIS Elements TB")
    iris["A1"] = "Account"
    iris["B1"] = "Description"
    iris["C1"] = ye_label
    style_header(iris, 1, 3)
    r = 2
    net = 0.0
    for account in sorted(buckets):
        amt = round(buckets[account], 2)
        if abs(amt) < 0.005:
            continue
        iris.cell(r, 1, account)
        iris.cell(r, 2, account)
        write_money(iris, r, 3, amt)
        net += amt
        r += 1
    iris.cell(r, 1, "Net (must be 0.00)")
    iris.cell(r, 1).font = Font(bold=True)
    write_money(iris, r, 3, net)
    iris.column_dimensions["A"].width = 62
    iris.column_dimensions["B"].width = 62
    iris.column_dimensions["C"].width = 18

    wb.save(dest)


def write_queries_md(path: Path, company: str, ye: str, queries: list[str], journals: list[dict]) -> None:
    lines = [
        f"# {company} — draft queries",
        f"",
        f"Year end {ye}. Draft only. Confirm these before IRIS finals / CH / HMRC.",
        "",
        "## Queries",
        "",
    ]
    for i, q in enumerate(queries, 1):
        lines.append(f"{i}. {q}")
    lines += ["", "## Draft journals posted in the pack", ""]
    for j in journals:
        side = "Dr" if j["amount"] >= 0 else "Cr"
        lines.append(f"- {j['jnl']} {side} {j['account']} £{abs(j['amount']):,.2f} — {j['narr']}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# Production Autos
# ---------------------------------------------------------------------------

AUTOS_MAP = {
    "Cash on hand": "Cash - Cash at bank and in hand",
    "Checking (8474)": "Cash - Cash at bank and in hand",
    "Debtors": "Debtors - Trade debtors",
    "Other debtors": "Debtors - Other debtors",
    "Uncategorised Asset": "Plant & machinery - Cost - additions",
    "Computer equipment additions at cost": "Computer equipment - Cost - additions",
    "Vehicles Purchased": "Motor vehicles - Cost - additions",
    "Creditors": "Creditors less than 1 year - Trade creditors",
    "Accruals": "Creditors less than 1 year - Accruals",
    "Corporation tax payable": "Creditors less than 1 year - Corporation tax",
    "Director's current account": "Creditors less than 1 year - Directors' loans",
    "Dividends payable": "Creditors less than 1 year - Proposed dividends",
    "Other loans": "Creditors less than 1 year - Finance lease and HP contracts",
    "Tax and National Insurance": "Creditors less than 1 year - Other taxes and social security",
    "VAT Control": "Creditors less than 1 year - Other taxes and social security",
    "VAT Suspense": "Creditors less than 1 year - Other taxes and social security",
    "Wages and salaries control": "Debtors - Other debtors",
    "Retained Earnings": "Profit and loss account - Brought forward",
    "Sales": "Turnover - Sales",
    "COS  Car Tax": "Cost of sales - Other direct costs",
    "CoS - Drivers": "Cost of sales - Subcontractor costs",
    "COS Car Hire": "Cost of sales - Hire of plant & machinery",
    "COS Insurance": "Cost of sales - Other direct costs",
    "COS Petrol": "Cost of sales - Other direct costs",
    "COS Repairs": "Cost of sales - Other direct costs",
    "Cost of sales": "Cost of sales - Purchases",
    "Advertising": "Administrative expenses - Legal & professional - Advertising and PR",
    "Bank charges": "Administrative expenses - General - Bank charges",
    "Computer running costs": "Administrative expenses - General - Software",
    "Directors' pension costs": "Administrative expenses - Employee costs - Pensions",
    "Directors' remuneration": "Administrative expenses - Employee costs - Directors' salaries",
    "Employer's NI contributions": "Administrative expenses - Employee costs - Employer's NI",
    "Insurance": "Administrative expenses - General - Insurance",
    "Legal and professional fees": "Administrative expenses - Legal & professional - Solicitors fees",
    "Light and heat": "Administrative expenses - Premises costs - Light and heat",
    "Office expenses, repairs & maintenance": "Administrative expenses - General - Repairs and maintenance",
    "Printing, postage and stationery": "Administrative expenses - General - Stationery and printing",
    "Rates": "Administrative expenses - Premises costs - Rates",
    "Rent": "Administrative expenses - Premises costs - Rent",
    "Subscriptions": "Administrative expenses - General - Subscriptions",
    "Repairs and maintenance": "Administrative expenses - General - Repairs and maintenance",
    "Telephone": "Administrative expenses - General - Telephone and fax",
    "Travelling expenses": "Administrative expenses - Employee costs - Travel and subsistence",
    "Dividend": "Profit and loss account - Equity dividends",
    "Uncategorised Expense": "Administrative expenses - General - Sundry expenses",
}


def load_autos_qbo() -> list[dict]:
    path = (
        CLIENTS
        / "Production Autos Limited"
        / "Current"
        / "Working Papers"
        / "Production Autos Limited - 28 Febrary 2026 - Year end accounts working papers.xlsm"
    )
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    ws = wb["2026QBTB"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=50, max_col=2, values_only=True):
        name = row[0]
        if not name or str(name).strip() in {"Quickbooks", ""}:
            continue
        amt = money(row[1])
        if abs(amt) < 0.005:
            continue
        src = str(name).strip()
        iris = AUTOS_MAP.get(src)
        if not iris:
            raise SystemExit(f"Unmapped Production Autos QBO account: {src!r}")
        rows.append({"source": src, "iris": iris, "amount": amt})
    wb.close()
    return rows


def prepare_autos(names: set[str]) -> dict:
    qbo = load_autos_qbo()
    buckets: dict[str, float] = defaultdict(float)
    mapping = []
    for row in qbo:
        buckets[row["iris"]] += row["amount"]
        mapping.append(
            {
                "source": row["source"],
                "iris": row["iris"],
                "qbo_amount": row["amount"],
                "journals": 0.0,
                "final": row["amount"],
            }
        )

    # Year-end journals that can be calculated from the existing registers.
    # Signs: debit positive, credit negative (IRIS convention).
    journals = [
        {"jnl": "Jnl1", "account": "Administrative expenses - General - Depreciation", "amount": 664.98, "narr": "Land & buildings (container) 15% NBV — OtherFA"},
        {"jnl": "Jnl1", "account": "Land & buildings - Depn - charge for the year", "amount": -664.98, "narr": "Land & buildings (container) 15% NBV — OtherFA"},
        {"jnl": "Jnl2", "account": "Administrative expenses - General - Depreciation", "amount": 3153.82, "narr": "Plant & machinery 20% NBV — OtherFA"},
        {"jnl": "Jnl2", "account": "Plant & machinery - Depn - charge for the year", "amount": -3153.82, "narr": "Plant & machinery 20% NBV — OtherFA"},
        {"jnl": "Jnl3", "account": "Administrative expenses - General - Depreciation", "amount": 336.50, "narr": "Computer equipment charge on old items only — OtherFA (2025/26 adds not depreciated)"},
        {"jnl": "Jnl3", "account": "Computer equipment - Depn - charge for the year", "amount": -336.50, "narr": "Computer equipment charge on old items only — OtherFA"},
        {"jnl": "Jnl5", "account": "Administrative expenses - General - Depreciation", "amount": 854.49, "narr": "Fixtures & fittings 20% NBV — OtherFA"},
        {"jnl": "Jnl5", "account": "Fixtures & fittings - Depn - charge for the year", "amount": -854.49, "narr": "Fixtures & fittings 20% NBV — OtherFA"},
        {"jnl": "Jnl6", "account": "Administrative expenses - General - Depreciation", "amount": 84565.23, "narr": "Motor vehicles charge per MV register (register does not agree to QBO Vehicles Purchased)"},
        {"jnl": "Jnl6", "account": "Motor vehicles - Depn - charge for the year", "amount": -84565.23, "narr": "Motor vehicles charge per MV register"},
        {"jnl": "Jnl7", "account": "Debtors - Other debtors", "amount": 76000.00, "narr": "Reclass P A Estates deposits 23–24 Oct 2025 (Uncategorised Expense)"},
        {"jnl": "Jnl7", "account": "Administrative expenses - General - Sundry expenses", "amount": -76000.00, "narr": "Reclass P A Estates deposits out of P&L"},
        {"jnl": "Jnl8", "account": "Administrative expenses - Legal & professional - Accountancy fees", "amount": 2000.00, "narr": "Draft accrual — CRM Accounts 2026-02-28 fee £2,000"},
        {"jnl": "Jnl8", "account": "Creditors less than 1 year - Accruals", "amount": -2000.00, "narr": "Draft accountancy accrual"},
    ]
    apply_jnls(buckets, journals)

    queries = [
        "Check company-car benefit (CAs) on the fleet — already flagged on the Notes tab. Several remaining cars have a list-price / CO2 figure in column B of MV.",
        "Motor vehicle register does not agree to QuickBooks Vehicles Purchased (£502,780.29 on the TB vs register cost carried forward £411,070.67). Additions in the year on QBO are £77,126.66; the register only lists MK08 XHF £14,000 and EU17 EHP £2,995. Need a complete buy/sell list for the year, including the Porsche proceeds (£61,000 banked 19 Sep 2025) while SV20 LGX is still on the register.",
        "Drawers £2,100 on OtherFA are marked RENT ??? — confirm capital furniture vs rent / Sign 2 Signs.",
        "P A Estates deposits £40,000 + £36,000 (23–24 Oct 2025) plus ITV unidentified £1,800. Draft-reclassified the £76,000 to other debtors. Confirm: rent deposit, property purchase, or director? Opening rent deposit on last year's papers was already £46,300.",
        "Wages and salaries control is a £355,248.43 DEBIT (£246,991.23 brought forward + £108,257.20 this year) and there is no 2026 wages expense on the QBO P&L. Last year statutory wages were £108,607. Do not file until payroll is recoded.",
        "Tax and NIC control is a £110,808.70 DEBIT (opened £71,490.29). VAT Control credit £11,602.18 vs VAT Suspense debit £153,699.14. These look like uncleared control accounts, not real debtors/creditors.",
        "Corporation tax payable is a £69,139.96 DEBIT on QBO (opened £68,667.57). Last year Accology had a CT creditor of £21,657. Need the 2025 CT computation and what was paid.",
        "Other loans is a £89,136.77 DEBIT (opened £49,140.48). Last year Accology had HP creditors £28,878 <1yr and £52,253 >1yr. Need HP statements and a <1yr / >1yr split.",
        "Cash on hand is a £155,936.56 CREDIT, unchanged year on year. Checking (8474) is overdrawn £39,162.84. Confirm both against bank statements.",
        "Accruals is a £615.62 DEBIT on QBO (wrong sign vs last year's Accology credit £1,881). Draft only adds the £2,000 accountancy accrual.",
        "Directors' loan: QBO credit £60,000. 2026 movements are Neil Adams deposits £26,000 + £2,000 + £10,000. Last year Accology DLA creditor was £82,938 with no opening breakdown — still needed.",
        "Trade debtors £73,703.20 and trade creditors £3,992.57 — TrDrs / TrCrs tabs in the pack are empty. Need aged listings.",
        "Computer register looks duplicated (Copeman £1,000 + £1,000 twice) and 2025/26 additions (including Tech Menders £697.99) have no depreciation. Charge posted is only £336.50 on the old items.",
        "No 2026 HP interest (last year £8,253), no staff wages, no employer's NIC. Profit is overstated until those are posted.",
        "Last year's Accology statutory FA / share capital £100 / deferred tax £13,366 are not on QBO. Overlay them before finals so the IRIS balance sheet continues from 28 Feb 2025.",
        "Dividends payable £5,000 debit sits on QBO both years — confirm whether a dividend was paid or this is a mispost.",
        "FD57 NKA is noted as owned by Cold Feet, managed by Production Autos, nil cost. Confirm not a company asset.",
    ]

    # P&L after journals
    def g(*keys: str) -> float:
        return round(sum(buckets.get(k, 0.0) for k in keys), 2)

    sales = -g("Turnover - Sales")
    purchases = g("Cost of sales - Purchases")
    subcon = g("Cost of sales - Subcontractor costs")
    hire = g("Cost of sales - Hire of plant & machinery")
    other_cos = g("Cost of sales - Other direct costs")
    cos = round(purchases + subcon + hire + other_cos, 2)
    gross = round(sales - cos, 2)
    depn = g("Administrative expenses - General - Depreciation")
    accountancy = g("Administrative expenses - Legal & professional - Accountancy fees")
    overheads_ex = round(
        g(
            "Administrative expenses - Employee costs - Directors' salaries",
            "Administrative expenses - Employee costs - Pensions",
            "Administrative expenses - Employee costs - Employer's NI",
            "Administrative expenses - Employee costs - Travel and subsistence",
            "Administrative expenses - Premises costs - Rent",
            "Administrative expenses - Premises costs - Rates",
            "Administrative expenses - Premises costs - Light and heat",
            "Administrative expenses - General - Telephone and fax",
            "Administrative expenses - General - Stationery and printing",
            "Administrative expenses - General - Bank charges",
            "Administrative expenses - General - Subscriptions",
            "Administrative expenses - General - Insurance",
            "Administrative expenses - General - Software",
            "Administrative expenses - General - Repairs and maintenance",
            "Administrative expenses - General - Sundry expenses",
            "Administrative expenses - Legal & professional - Solicitors fees",
            "Administrative expenses - Legal & professional - Advertising and PR",
        ),
        2,
    )
    overheads = round(overheads_ex + depn + accountancy, 2)
    profit = round(gross - overheads, 2)

    fa_nbv = round(
        g(
            "Land & buildings - Cost - b/fwd",
            "Land & buildings - Cost - additions",
            "Land & buildings - Depn - b/fwd",
            "Land & buildings - Depn - charge for the year",
            "Plant & machinery - Cost - b/fwd",
            "Plant & machinery - Cost - additions",
            "Plant & machinery - Depn - b/fwd",
            "Plant & machinery - Depn - charge for the year",
            "Motor vehicles - Cost - b/fwd",
            "Motor vehicles - Cost - additions",
            "Motor vehicles - Depn - b/fwd",
            "Motor vehicles - Depn - charge for the year",
            "Fixtures & fittings - Cost - b/fwd",
            "Fixtures & fittings - Cost - additions",
            "Fixtures & fittings - Depn - b/fwd",
            "Fixtures & fittings - Depn - charge for the year",
            "Computer equipment - Cost - b/fwd",
            "Computer equipment - Cost - additions",
            "Computer equipment - Depn - b/fwd",
            "Computer equipment - Depn - charge for the year",
        ),
        2,
    )
    debtors = g("Debtors - Trade debtors", "Debtors - Other debtors", "Debtors - Accrued income and prepayments")
    cash = g("Cash - Cash at bank and in hand")
    ca = round(debtors + cash, 2)
    creditors = -g(
        "Creditors less than 1 year - Trade creditors",
        "Creditors less than 1 year - Accruals",
        "Creditors less than 1 year - Corporation tax",
        "Creditors less than 1 year - Directors' loans",
        "Creditors less than 1 year - Proposed dividends",
        "Creditors less than 1 year - Finance lease and HP contracts",
        "Creditors less than 1 year - Other taxes and social security",
        "Creditors less than 1 year - Other creditors",
        "Creditors greater than 1 year - Finance lease and HP contracts",
    )
    net_assets = round(fa_nbv + ca - creditors, 2)
    equity = -g(
        "Share capital - Brought forward",
        "Profit and loss account - Brought forward",
        "Profit and loss account - Equity dividends",
    )
    # current year profit sits in P&L lines, so funds = equity + profit
    funds = round(equity + profit, 2)

    client_dir = CLIENTS / "Production Autos Limited"
    pack = client_dir / "Current" / "Working Papers" / "Production Autos Limited - 28 February 2026 - Draft accounts pack.xlsx"
    iris = client_dir / "Current" / "IRIS Import" / "2026-02-28 IRIS Elements TB.csv"
    map_csv = client_dir / "Current" / "IRIS Import" / "2026-02-28 Accology chart mapping.csv"
    qmd = client_dir / "Current" / "Working Papers" / "Production Autos Limited - 28 February 2026 - Queries.md"

    net, unknown = write_iris_csv(iris, "Year End 28/02/2026", buckets, names)
    write_mapping_csv(map_csv, mapping)
    write_queries_md(qmd, "Production Autos Limited", "28 February 2026", queries, journals)
    write_pack(
        pack,
        company="Production Autos Limited",
        subtitle="Draft accounts working pack — year ended 28 February 2026",
        source_note=(
            "Prepared from Current/Working Papers/"
            "Production Autos Limited - 28 Febrary 2026 - Year end accounts working papers.xlsm "
            "(QBO TB as at 28 Feb 2026, already mapped to Accology names on PL/BS)."
        ),
        ye_label="Year ended 28 February 2026",
        qbo_rows=qbo,
        journals=journals,
        buckets=buckets,
        queries=queries,
        cover_stats=[
            ("QBO TB net (should be 0.00)", 0.0),
            ("Draft IRIS TB net (must be 0.00)", net),
            ("Turnover", sales),
            ("Gross profit", gross),
            ("Depreciation charged (draft)", depn),
            ("Draft profit before tax", profit),
            ("Net assets (draft, QBO + journals)", net_assets),
            ("Shareholders' funds check", funds),
        ],
        pl_lines=[
            ("Turnover", sales),
            ("Purchases", -purchases),
            ("Drivers / subcontractors", -subcon),
            ("Car hire", -hire),
            ("Other direct costs (tax, petrol, repairs)", -other_cos),
            ("Cost of sales", -cos),
            ("Gross profit", gross),
            ("Directors' remuneration and pensions", -g("Administrative expenses - Employee costs - Directors' salaries", "Administrative expenses - Employee costs - Pensions")),
            ("Premises (rent, rates, heat)", -g("Administrative expenses - Premises costs - Rent", "Administrative expenses - Premises costs - Rates", "Administrative expenses - Premises costs - Light and heat")),
            ("Other overheads", -round(overheads_ex - g("Administrative expenses - Employee costs - Directors' salaries", "Administrative expenses - Employee costs - Pensions", "Administrative expenses - Premises costs - Rent", "Administrative expenses - Premises costs - Rates", "Administrative expenses - Premises costs - Light and heat"), 2)),
            ("Accountancy accrual (draft)", -accountancy),
            ("Depreciation (draft)", -depn),
            ("Profit before tax", profit),
            ("Tax (not computed — see Queries)", 0.0),
            ("Profit for the year (before tax)", profit),
        ],
        bs_sections=[
            (
                "Fixed assets (QBO cost + draft depreciation only)",
                [
                    ("Plant / computers / vehicles at QBO cost, less draft depn", fa_nbv),
                    ("Total fixed assets", fa_nbv),
                ],
            ),
            (
                "Current assets",
                [
                    ("Trade debtors", g("Debtors - Trade debtors")),
                    ("Other debtors (includes wages control + PA Estates deposits)", g("Debtors - Other debtors")),
                    ("Cash at bank and in hand (includes cash-on-hand credit)", cash),
                    ("Total current assets", ca),
                ],
            ),
            (
                "Creditors",
                [
                    ("Trade creditors", -g("Creditors less than 1 year - Trade creditors")),
                    ("Accruals (QBO debit + £2,000 draft)", -g("Creditors less than 1 year - Accruals")),
                    ("Corporation tax (QBO is a debit — query)", -g("Creditors less than 1 year - Corporation tax")),
                    ("Directors' loans", -g("Creditors less than 1 year - Directors' loans")),
                    ("Proposed dividends (QBO debit — query)", -g("Creditors less than 1 year - Proposed dividends")),
                    ("HP / other loans (QBO is a debit — query)", -g("Creditors less than 1 year - Finance lease and HP contracts")),
                    ("VAT / PAYE / NIC controls (net — query)", -g("Creditors less than 1 year - Other taxes and social security")),
                    ("Total creditors as mapped", creditors),
                ],
            ),
            (
                "Capital and reserves",
                [
                    ("Profit and loss brought forward (QBO retained earnings)", -g("Profit and loss account - Brought forward")),
                    ("Profit for the year (before tax)", profit),
                    ("Shareholders' funds (draft)", funds),
                    ("Net assets (draft)", net_assets),
                ],
            ),
        ],
    )
    return {
        "company": "Production Autos Limited",
        "pack": pack,
        "iris": iris,
        "queries": qmd,
        "net": net,
        "unknown": unknown,
        "profit": profit,
        "n_q": len(queries),
        "n_j": len({j["jnl"] for j in journals}),
    }


# ---------------------------------------------------------------------------
# Big Box Art
# ---------------------------------------------------------------------------

BBA_MAP = {
    "Adam Personal Natwest": "Cash - Cash at bank and in hand",
    "Amazon Seller Cent. - Clearing Account": "Cash - Cash at bank and in hand",
    "Amazon Vendor Cent. - Clearing Account": "Cash - Cash at bank and in hand",
    "Barclays Business UK": "Cash - Cash at bank and in hand",
    "Belmont 6 Website - Clearing Account": "Cash - Cash at bank and in hand",
    "CO-OP Bank": "Cash - Cash at bank and in hand",
    "eBay - Clearing Account": "Cash - Cash at bank and in hand",
    "HSBC Business": "Cash - Cash at bank and in hand",
    "Paypal UK": "Cash - Cash at bank and in hand",
    "Rebecca Personal Bank": "Cash - Cash at bank and in hand",
    "Savings Account - Barclays": "Cash - Cash at bank and in hand",
    "Debtors": "Debtors - Trade debtors",
    "Computer Equipment  Accumulated Depreciation": "Computer equipment - Depn - b/fwd",
    "Computer equipment cost brought forward": "Computer equipment - Cost - b/fwd",
    "Goodwill": "Intangible Fixed Assets - Goodwill - Cost - b/fwd",
    "Goodwill Amortisation": "Intangible Fixed Assets - Goodwill - Amortisation - b/fwd",
    "Office Equipment Accumulated Depreciation": "Fixtures & fittings - Depn - b/fwd",
    "Office furniture fittings equipment cost brought forward": "Fixtures & fittings - Cost - b/fwd",
    "Creditors": "Creditors less than 1 year - Trade creditors",
    "Accruals": "Creditors less than 1 year - Accruals",
    "Corporation tax payable": "Creditors less than 1 year - Corporation tax",
    "Payroll Liabilities:HMRC": "Creditors less than 1 year - Other taxes and social security",
    "VAT Control": "Creditors less than 1 year - Other taxes and social security",
    "VAT Suspense": "Creditors less than 1 year - Other taxes and social security",
    "Adam BG Drawings": "Creditors less than 1 year - Directors' loans",
    "Opening Balance Equity": "Profit and loss account - Brought forward",
    "Rebecca BG Drawings": "Creditors less than 1 year - Directors' loans",
    "Retained Earnings": "Profit and loss account - Brought forward",
    "Amazon Prime Sales": "Turnover - Sales",
    "Amazon Seller Central Sales": "Turnover - Sales",
    "Customer Refunds:Amazon SC Refunds": "Turnover - Sales",
    "Customer Refunds:eBay Refunds": "Turnover - Sales",
    "Customer Refunds:Wayfair Refunds": "Turnover - Sales",
    "Discounts & Allowances:Wayfair Discounts & Allowances": "Cost of sales - Discounts allowed",
    "eBay Sales": "Turnover - Sales",
    "Sales of Product Income": "Turnover - Sales",
    "Wayfair Sales (Big Box Art)": "Turnover - Sales",
    "Drop Ship Supply Purchases": "Cost of sales - Purchases",
    "Advertising": "Administrative expenses - Legal & professional - Advertising and PR",
    "Amazon Customer Refunds": "Cost of sales - Purchases",
    "Amazon Fees": "Cost of sales - Other direct costs",
    "Amazon Receivables": "Cost of sales - Other direct costs",
    "Bank charges": "Administrative expenses - General - Bank charges",
    "Charitable donations": "Administrative expenses - General - Donations",
    "Cloud Services": "Administrative expenses - General - Software",
    "Computer Hardware": "Computer equipment - Cost - additions",
    "Courier and delivery charges": "Cost of sales - Carriage",
    "eBay Fees": "Cost of sales - Other direct costs",
    "Ebay Payment Holds": "Cost of sales - Other direct costs",
    "Etsy Fees": "Cost of sales - Other direct costs",
    "HMRC Taxes Paid": "Debtors - Other debtors",
    "Insurance": "Administrative expenses - General - Insurance",
    "Legal and professional fees": "Administrative expenses - Legal & professional - Other legal and professional",
    "Office Equipment": "Fixtures & fittings - Cost - additions",
    "Office expenses, repairs & maintenance": "Administrative expenses - General - Repairs and maintenance",
    "Payroll Expenses:Taxes": "Administrative expenses - Employee costs - Employer's NI",
    "Payroll Expenses:Wages": "Administrative expenses - Employee costs - Wages and salaries",
    "Product Marketing Materials": "Administrative expenses - Legal & professional - Advertising and PR",
    "Rent": "Administrative expenses - Premises costs - Rent",
    "Software Purchases": "Administrative expenses - General - Software",
    "Software Subscriptions": "Administrative expenses - General - Software",
    "Staff Meals (Travel)": "Administrative expenses - Employee costs - Travel and subsistence",
    "Staff training": "Administrative expenses - Employee costs - Staff training and welfare",
    "Subscriptions": "Administrative expenses - General - Subscriptions",
    "Sundry expenses": "Administrative expenses - General - Sundry expenses",
    "Wages": "Administrative expenses - Employee costs - Wages and salaries",
    "Website Operating Costs": "Administrative expenses - General - Software",
    "Savings Account Interest earned": "Interest receivable",
    "Dividend paid": "Profit and loss account - Equity dividends",
}


def load_bba_qbo() -> list[dict]:
    path = (
        CLIENTS
        / "Big Box Art Limited"
        / "Accounts"
        / "Big Box Art Limited - 31 December 2025 - Year end Accounts Wokring Papers.xlsm"
    )
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    ws = wb["QBTB"]
    rows = []
    for row in ws.iter_rows(min_row=6, max_row=74, max_col=3, values_only=True):
        name = row[0]
        if not name or str(name).strip().upper() == "TOTAL":
            continue
        amt = money(row[1]) - money(row[2])
        if abs(amt) < 0.005:
            continue
        src = str(name).strip()
        iris = BBA_MAP.get(src)
        if not iris:
            raise SystemExit(f"Unmapped Big Box Art QBO account: {src!r}")
        rows.append({"source": src, "iris": iris, "amount": amt})
    wb.close()
    return rows


def prepare_bba(names: set[str]) -> dict:
    qbo = load_bba_qbo()
    buckets: dict[str, float] = defaultdict(float)
    mapping = []
    for row in qbo:
        buckets[row["iris"]] += row["amount"]
        mapping.append(
            {
                "source": row["source"],
                "iris": row["iris"],
                "qbo_amount": row["amount"],
                "journals": 0.0,
                "final": row["amount"],
            }
        )

    # Prior-year FA journals from the 2024 WTB that were never posted back to QBO.
    # P&L impact goes to reserves, not 2025 profit.
    journals = [
        {"jnl": "PY1", "account": "Profit and loss account - Brought forward", "amount": 20000.00, "narr": "2023+2024 goodwill amortisation still missing from QBO (10 years, £10k p.a.)"},
        {"jnl": "PY1", "account": "Intangible Fixed Assets - Goodwill - Amortisation - b/fwd", "amount": -20000.00, "narr": "Bring QBO amortisation from £20k to 2024 WTB £40k"},
        {"jnl": "PY2", "account": "Computer equipment - Cost - b/fwd", "amount": 4088.07, "narr": "2023 £639.36 + 2024 £3,448.71 computer additions capitalised on 2024 WTB, not in QBO"},
        {"jnl": "PY2", "account": "Profit and loss account - Brought forward", "amount": -4088.07, "narr": "PY computer cost overlay"},
        {"jnl": "PY3", "account": "Profit and loss account - Brought forward", "amount": 2566.26, "narr": "2023 £1,297 + 2024 £1,269.26 computer depn on 2024 WTB, not in QBO"},
        {"jnl": "PY3", "account": "Computer equipment - Depn - b/fwd", "amount": -2566.26, "narr": "PY computer depn overlay"},
        {"jnl": "PY4", "account": "Fixtures & fittings - Cost - b/fwd", "amount": 1231.18, "narr": "2023 fixtures additions on 2024 WTB, not in QBO"},
        {"jnl": "PY4", "account": "Profit and loss account - Brought forward", "amount": -1231.18, "narr": "PY fixtures cost overlay"},
        {"jnl": "PY5", "account": "Profit and loss account - Brought forward", "amount": 1008.92, "narr": "2023 £545 + 2024 £463.92 fixtures depn on 2024 WTB, not in QBO"},
        {"jnl": "PY5", "account": "Fixtures & fittings - Depn - b/fwd", "amount": -1008.92, "narr": "PY fixtures depn overlay"},
        # 2025 current-year journals
        {"jnl": "Jnl1", "account": "Administrative expenses - General - Amortisation of goodwill", "amount": 10000.00, "narr": "2025 goodwill amortisation — 10-year policy per 2024 WTB"},
        {"jnl": "Jnl1", "account": "Intangible Fixed Assets - Goodwill - Amortisation - provided in year", "amount": -10000.00, "narr": "2025 goodwill amortisation"},
        {"jnl": "Jnl2", "account": "Administrative expenses - General - Depreciation", "amount": 1564.85, "narr": "Computers: remaining 1/3 of 2024 adds £1,149.57 + 1/3 of 2025 adds £415.28 (same rule as 2024 WTB)"},
        {"jnl": "Jnl2", "account": "Computer equipment - Depn - charge for the year", "amount": -1564.85, "narr": "2025 computer depreciation"},
        {"jnl": "Jnl3", "account": "Administrative expenses - General - Depreciation", "amount": 895.34, "narr": "Fixtures 25% NBV on 2024 WTB NBV £2,624.00 + 2025 additions £957.36"},
        {"jnl": "Jnl3", "account": "Fixtures & fittings - Depn - charge for the year", "amount": -895.34, "narr": "2025 fixtures depreciation"},
        {"jnl": "Jnl4", "account": "Administrative expenses - Legal & professional - Accountancy fees", "amount": 3000.00, "narr": "Draft accrual — CRM Accounts 2025-12-31 fee £3,000"},
        {"jnl": "Jnl4", "account": "Creditors less than 1 year - Accruals", "amount": -3000.00, "narr": "Draft accountancy accrual (old £2,400 accrual still on QBO — query)"},
    ]
    apply_jnls(buckets, journals)

    queries = [
        "Trade creditors £97,752.93. Trade Canvas Co. £74,514.02 (of which £40,897.71 is 31–60 days and £6,348.31 is 91+). Yoosh £19,031.64. Confirm still payable and whether any stock / credits are missing.",
        "HMRC Taxes Paid £9,006.22 is a P&L code on QBO. Draft parks it in other debtors. 2024 analysis split this between directors' self assessment (to DLA) and corporation tax. Need the 2025 payment list coded the same way.",
        "Corporation tax creditor £10,810 is unchanged on QBO since 2023. 2024 WTB cleared it to nil after payments through HMRC Taxes Paid. Those journals were never posted back to QBO. Confirm 2024 CT paid and compute 2025 CT.",
        "Opening Balance Equity £138,539.10 credit is still on QBO. 2024 WTB transferred it to retained earnings. Draft maps it to P&L reserve. Confirm share capital (not on QBO).",
        "Adam BG Drawings credit £55,488.06 and Rebecca BG Drawings credit £54,981.73 mapped to directors' loans (same as 2024). Need DLA movements for 2025 (dividends already sit in Dividend paid £50,049.95).",
        "Payroll Liabilities HMRC credit £2,601.38 vs wages £25,240.05 + employer's NIC £668.60. Confirm RTI / FPS and whether a year-end PAYE creditor is still due.",
        "VAT Control credit £2.49 + VAT Suspense credit £3,506.27. VAT 100 for the calendar year: Box 5 net £17,951.10. These do not agree — need VAT returns vs QBO.",
        "Computer Hardware £1,245.83 and Office Equipment £957.36 capitalised (same treatment as 2024 computers). Confirm nothing here is a repair.",
        "Goodwill still being amortised at £10,000 a year (10 years from 2021). After PY overlay + 2025 charge, cost £100,000 / amort £50,000. Confirm policy still applies.",
        "Old accountancy accrual £2,400 remains on QBO. Draft adds a further £3,000. Confirm whether 2024 was billed and if £2,400 should be released.",
        "Trade debtors £0.01 (Wayfair). Sales £590,940 — confirm marketplaces are fully settled and no Amazon/eBay holds belong in debtors (Ebay Payment Holds £566.13 left in COS).",
        "Adam Personal Natwest is a £168.80 credit (overdrawn personal). 2024 transferred this to DLA. Left in cash on the draft — confirm.",
        "No stock on the TB (drop-ship). Confirm nothing is held at the year end.",
        "2023 and 2024 year-end journals were never posted back to QuickBooks. Draft overlays the 2024 WTB FA numbers. Post those journals to QBO so next year starts clean.",
        "Legal and professional £3,435 — split accountancy vs other if any of this is the 2024 Accology fee already accrued.",
        "Donations £53.75 — add back for CT unless a qualifying charity with evidence.",
    ]

    def g(*keys: str) -> float:
        return round(sum(buckets.get(k, 0.0) for k in keys), 2)

    sales = -g("Turnover - Sales")
    purchases = g("Cost of sales - Purchases")
    carriage = g("Cost of sales - Carriage")
    discounts = g("Cost of sales - Discounts allowed")
    other_cos = g("Cost of sales - Other direct costs")
    cos = round(purchases + carriage + discounts + other_cos, 2)
    gross = round(sales - cos, 2)
    wages = g("Administrative expenses - Employee costs - Wages and salaries")
    nic = g("Administrative expenses - Employee costs - Employer's NI")
    software = g("Administrative expenses - General - Software")
    ads = g("Administrative expenses - Legal & professional - Advertising and PR")
    legal = g("Administrative expenses - Legal & professional - Other legal and professional")
    accountancy = g("Administrative expenses - Legal & professional - Accountancy fees")
    other_oh = g(
        "Administrative expenses - General - Bank charges",
        "Administrative expenses - General - Donations",
        "Administrative expenses - General - Insurance",
        "Administrative expenses - General - Repairs and maintenance",
        "Administrative expenses - General - Subscriptions",
        "Administrative expenses - General - Sundry expenses",
        "Administrative expenses - Premises costs - Rent",
        "Administrative expenses - Employee costs - Travel and subsistence",
        "Administrative expenses - Employee costs - Staff training and welfare",
    )
    depn = g("Administrative expenses - General - Depreciation")
    amort = g("Administrative expenses - General - Amortisation of goodwill")
    interest = -g("Interest receivable")
    overheads = round(wages + nic + software + ads + legal + accountancy + other_oh + depn + amort, 2)
    profit = round(gross - overheads + interest, 2)
    div = g("Profit and loss account - Equity dividends")

    fa = round(
        g(
            "Intangible Fixed Assets - Goodwill - Cost - b/fwd",
            "Intangible Fixed Assets - Goodwill - Amortisation - b/fwd",
            "Intangible Fixed Assets - Goodwill - Amortisation - provided in year",
            "Computer equipment - Cost - b/fwd",
            "Computer equipment - Cost - additions",
            "Computer equipment - Depn - b/fwd",
            "Computer equipment - Depn - charge for the year",
            "Fixtures & fittings - Cost - b/fwd",
            "Fixtures & fittings - Cost - additions",
            "Fixtures & fittings - Depn - b/fwd",
            "Fixtures & fittings - Depn - charge for the year",
        ),
        2,
    )
    debtors = g("Debtors - Trade debtors", "Debtors - Other debtors")
    cash = g("Cash - Cash at bank and in hand")
    ca = round(debtors + cash, 2)
    creditors = -g(
        "Creditors less than 1 year - Trade creditors",
        "Creditors less than 1 year - Accruals",
        "Creditors less than 1 year - Corporation tax",
        "Creditors less than 1 year - Other taxes and social security",
        "Creditors less than 1 year - Directors' loans",
    )
    net_assets = round(fa + ca - creditors, 2)
    reserves = -g("Profit and loss account - Brought forward", "Profit and loss account - Equity dividends")
    funds = round(reserves + profit, 2)

    client_dir = CLIENTS / "Big Box Art Limited"
    wp = client_dir / "Current" / "Working Papers"
    iris_dir = client_dir / "Current" / "IRIS Import"
    pack = wp / "Big Box Art Limited - 31 December 2025 - Draft accounts pack.xlsx"
    iris = iris_dir / "2025-12-31 IRIS Elements TB.csv"
    map_csv = iris_dir / "2025-12-31 Accology chart mapping.csv"
    qmd = wp / "Big Box Art Limited - 31 December 2025 - Queries.md"

    net, unknown = write_iris_csv(iris, "Year End 31/12/2025", buckets, names)
    write_mapping_csv(map_csv, mapping)
    write_queries_md(qmd, "Big Box Art Limited", "31 December 2025", queries, journals)
    write_pack(
        pack,
        company="Big Box Art Limited",
        subtitle="Draft accounts working pack — year ended 31 December 2025",
        source_note=(
            "Prepared from Accounts/"
            "Big Box Art Limited - 31 December 2025 - Year end Accounts Wokring Papers.xlsm "
            "(QBO TB as at 31 Dec 2025) plus the 2024 WTB FA journals that were never posted back to QBO."
        ),
        ye_label="Year ended 31 December 2025",
        qbo_rows=qbo,
        journals=journals,
        buckets=buckets,
        queries=queries,
        cover_stats=[
            ("QBO TB balanced", 1011028.25),
            ("Draft IRIS TB net (must be 0.00)", net),
            ("Turnover (net of refunds)", sales),
            ("Gross profit", gross),
            ("Draft profit before tax", profit),
            ("Dividends", div),
            ("Net assets (draft)", net_assets),
            ("Shareholders' funds check", funds),
        ],
        pl_lines=[
            ("Turnover (sales less marketplace refunds)", sales),
            ("Purchases / drop-ship / Amazon customer refunds", -purchases),
            ("Carriage", -carriage),
            ("Discounts allowed", -discounts),
            ("Marketplace fees and other direct costs", -other_cos),
            ("Cost of sales", -cos),
            ("Gross profit", gross),
            ("Wages", -wages),
            ("Employer's NIC", -nic),
            ("Software / website / cloud", -software),
            ("Advertising and marketing", -ads),
            ("Legal and professional", -legal),
            ("Accountancy accrual (draft)", -accountancy),
            ("Other overheads", -other_oh),
            ("Depreciation (draft)", -depn),
            ("Goodwill amortisation (draft)", -amort),
            ("Interest receivable", interest),
            ("Profit before tax", profit),
            ("Tax (not computed — see Queries)", 0.0),
            ("Profit for the year (before tax)", profit),
            ("Equity dividends", -div),
        ],
        bs_sections=[
            (
                "Fixed assets",
                [
                    ("Goodwill (cost £100,000 less amort £50,000)", round(g("Intangible Fixed Assets - Goodwill - Cost - b/fwd", "Intangible Fixed Assets - Goodwill - Amortisation - b/fwd", "Intangible Fixed Assets - Goodwill - Amortisation - provided in year"), 2)),
                    ("Computer equipment", round(g("Computer equipment - Cost - b/fwd", "Computer equipment - Cost - additions", "Computer equipment - Depn - b/fwd", "Computer equipment - Depn - charge for the year"), 2)),
                    ("Fixtures & fittings", round(g("Fixtures & fittings - Cost - b/fwd", "Fixtures & fittings - Cost - additions", "Fixtures & fittings - Depn - b/fwd", "Fixtures & fittings - Depn - charge for the year"), 2)),
                    ("Total fixed assets", fa),
                ],
            ),
            (
                "Current assets",
                [
                    ("Trade debtors", g("Debtors - Trade debtors")),
                    ("Other debtors (HMRC Taxes Paid parked)", g("Debtors - Other debtors")),
                    ("Cash at bank and in hand", cash),
                    ("Total current assets", ca),
                ],
            ),
            (
                "Creditors: amounts falling due within one year",
                [
                    ("Trade creditors", -g("Creditors less than 1 year - Trade creditors")),
                    ("Accruals (old £2,400 + draft £3,000)", -g("Creditors less than 1 year - Accruals")),
                    ("Corporation tax (stale QBO balance — query)", -g("Creditors less than 1 year - Corporation tax")),
                    ("VAT / PAYE", -g("Creditors less than 1 year - Other taxes and social security")),
                    ("Directors' loans (drawings credits)", -g("Creditors less than 1 year - Directors' loans")),
                    ("Total creditors", creditors),
                ],
            ),
            (
                "Capital and reserves",
                [
                    ("Profit and loss brought forward (QBO RE + OBE + PY overlays)", -g("Profit and loss account - Brought forward")),
                    ("Profit for the year (before tax)", profit),
                    ("Equity dividends", -div),
                    ("Shareholders' funds (draft)", funds),
                    ("Net assets (draft)", net_assets),
                ],
            ),
        ],
    )
    return {
        "company": "Big Box Art Limited",
        "pack": pack,
        "iris": iris,
        "queries": qmd,
        "net": net,
        "unknown": unknown,
        "profit": profit,
        "n_q": len(queries),
        "n_j": len({j["jnl"] for j in journals}),
    }


def main() -> int:
    names = chart_names()
    autos = prepare_autos(names)
    bba = prepare_bba(names)
    rc = 0
    for res in (autos, bba):
        print(f"{res['company']}: pack={res['pack']}")
        print(f"  IRIS={res['iris']}  net={res['net']}  profit={res['profit']}")
        print(f"  queries={res['n_q']} journals={res['n_j']}  md={res['queries']}")
        if res["unknown"]:
            print("  UNMAPPED", res["unknown"])
            rc = 1
        if abs(res["net"]) > 0.02:
            print("  TB OUT OF BALANCE")
            rc = 1
    return rc


if __name__ == "__main__":
    raise SystemExit(main())
