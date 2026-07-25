"""Restore opening-balance invoice dates from Aged Receivables Detail xlsx."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models.sales import Invoice

SOURCE = Path(
    r"C:\Users\SimonDuckworth\Downloads\Accology_Limited_-_Aged_Receivables_Detail.xlsx"
)


def load_dates_from_xlsx(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_num: dict = {}
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not cells:
            continue
        if cells[0] == "Invoice Date" and len(cells) > 2 and cells[2] == "Invoice Number":
            header_seen = True
            continue
        if not header_seen:
            continue
        if len(cells) < 3 or cells[2] is None:
            continue
        inv_num = str(cells[2]).strip()
        if not inv_num.upper().startswith("INV"):
            continue
        inv_date = cells[0]
        due_date = cells[1] if len(cells) > 1 else None
        if isinstance(inv_date, datetime):
            inv_date = inv_date.date()
        if isinstance(due_date, datetime):
            due_date = due_date.date()
        if not isinstance(inv_date, date):
            continue
        by_num[inv_num] = {
            "issue": inv_date,
            "due": due_date if isinstance(due_date, date) else inv_date,
        }
    wb.close()
    return by_num


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Source not found: {SOURCE}")
    by_num = load_dates_from_xlsx(SOURCE)
    print(f"parsed {len(by_num)} invoices from {SOURCE.name}")

    db = SessionLocal()
    invs = db.query(Invoice).filter(Invoice.source == "opening_balance").all()
    updated = 0
    missing = []
    for inv in invs:
        info = by_num.get(inv.number or "")
        if not info:
            missing.append(inv.number)
            continue
        inv.issue_date = info["issue"]
        inv.due_date = info["due"]
        inv.notes = (
            f"Opening balance · invoice date {info['issue'].isoformat()} "
            f"(restored from Aged Receivables Detail)"
        )
        updated += 1
    db.commit()
    print(f"updated {updated}")
    if missing:
        print(f"missing from source ({len(missing)}): {missing}")

    invs = db.query(Invoice).filter(Invoice.source == "opening_balance").all()
    still_today = sum(1 for i in invs if str(i.issue_date) == "2026-07-25")
    print(f"still today: {still_today}")
    print("sample:", [(i.number, str(i.issue_date), str(i.due_date)) for i in invs[:10]])
    db.close()


if __name__ == "__main__":
    main()
