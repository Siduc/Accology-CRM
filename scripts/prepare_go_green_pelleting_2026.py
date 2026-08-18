"""Prepare Go Green Pelleting Solutions YE 31/03/2026 pack from Mike's TB.

Does not submit to IRIS. Writes working papers + IRIS-ready recoded TB.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.database import SessionLocal
from app.models.client import Client
from app.services.client_playbook import set_bookkeeping_source

SRC = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Go Green Pelleting Solutions Limited\Current\Source"
    r"\Go-Green-Pelleting-Solutions-2026-accounts.xlsx"
)
CLIENT_DIR = Path(
    r"C:\Users\User\OneDrive - Accology\Accologise Documents\Clients"
    r"\Go Green Pelleting Solutions Limited"
)

MIKE_IDS = (
    283,
    119,
    76,
    67,
    37,
    65,
    60,
    28,
    125,
    58,
    80,
    108,
    111,
    57,
    118,
)

NOTE = (
    "Anything from Mike Jacques comes as a ready trial balance by email "
    "(Outlook Holding). Recode to IRIS chart, send drafts for approval, "
    "amend, submit through IRIS, send finals. No Xero/Sage/QBO pull or post-back."
)

NAVY = "052891"
GOLD = "F4B809"
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def money(n) -> float:
    try:
        return round(float(n or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def parse_tb(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(11, 83):
        name = ws[f"C{r}"].value
        if not name or not str(name).strip():
            continue
        dr = money(ws[f"I{r}"].value)
        cr = money(ws[f"K{r}"].value)
        rows.append({"row": r, "name": str(name).strip(), "dr": dr, "cr": cr})
    tot_dr = money(ws["I84"].value)
    tot_cr = money(ws["K84"].value)
    return rows, tot_dr, tot_cr


def nvl(rows, *names, side="dr"):
    total = 0.0
    for row in rows:
        if row["name"] in names:
            total += row[side]
    return round(total, 2)


def build_pack(rows, tot_dr, tot_cr) -> dict:
    # Closing FA
    plant_cost = nvl(rows, "Plant and machinery cost brought forward") + nvl(
        rows, "Plant and machinery additions"
    ) + nvl(rows, "Plant and machinery transfer to ownership cost", side="dr")
    plant_dep = (
        nvl(rows, "Plant and machinery depreciation brought forward", side="cr")
        + nvl(rows, "Plant and machinery transfer to ownership depreciation", side="cr")
        + nvl(rows, "Plant and machinery charge for the year", side="cr")
    )
    lease_p_cost = nvl(rows, "Plant and machinery lease cost additions")
    lease_p_dep = nvl(
        rows, "Plant and mnachinery lease depreciation charge for the year", side="cr"
    )
    ff_cost = nvl(rows, "Fixtures and fittings cost brought forward")
    ff_dep = nvl(rows, "Fixtures and fittings depreciation brought forward", side="cr") + nvl(
        rows, "Fixtures and fittings depreciation charge for the year", side="cr"
    )
    mv_cost = nvl(rows, "Motor vehicles transfer to ownership cost") - nvl(
        rows, "Motor vehicles disposal cost", side="cr"
    )
    mv_dep = (
        nvl(rows, "Motor vehicles transfer to ownership depreciation", side="cr")
        - nvl(rows, "Motor vehicles depreciation on disposal")
        + nvl(rows, "Motor vehicles depreciation charge for the year", side="cr")
    )
    lease_mv_cost = nvl(rows, "Motor vehicles lease cost brought forward") - nvl(
        rows, "Motor vehicles lease cost transfer to ownership", side="cr"
    )
    lease_mv_dep = (
        nvl(rows, "Motor vehicles lease depreciation brought forward", side="cr")
        - nvl(rows, "Motor vehicles transfer to ownership depreciation")
        + nvl(rows, "Motor vehicles lease depreciation charge for the year", side="cr")
    )
    it_cost = nvl(rows, "Computer equipment cost brought forward")
    it_dep = nvl(rows, "Computer equipmemt depreciation brought forward", side="cr") + nvl(
        rows, "Computer equipmemt depreciation charge for the year", side="cr"
    )

    debtors = nvl(rows, "Other debtors")
    bank = nvl(rows, "Bank account")
    vat = nvl(rows, "VAT")
    kelly_cr = nvl(rows, "D Kelly", side="cr")
    swift_cr = nvl(rows, "M Swift", side="cr")
    paye = nvl(rows, "PAYE", side="cr")
    accruals = nvl(rows, "Accruals", side="cr")
    oth_cr = nvl(rows, "Other creditors", side="cr")
    loan = nvl(rows, "Bank loan", side="cr")
    hp = sum(r["cr"] for r in rows if r["name"] == "HP")
    hp = round(hp, 2)
    dtax = nvl(rows, "Deferred tax", side="cr")
    shares = nvl(rows, "Share capital", side="cr")
    pl_bf = nvl(rows, "Profit and loss account", side="cr")
    div = nvl(rows, "Dividends")

    sales = nvl(rows, "Sales", side="cr")
    purchases = nvl(rows, "Purchases")
    haulage = nvl(rows, "Haulage")
    insurance = nvl(rows, "Insurance")
    kelly_pay = sum(r["dr"] for r in rows if r["name"] == "D Kelly")
    swift_pay = sum(r["dr"] for r in rows if r["name"] == "M Swift")
    wages = nvl(rows, "Wages")
    nic = nvl(rows, "Social security")
    admin = sum(
        nvl(rows, n)
        for n in (
            "Telephone",
            "Post and stationery",
            "Travelling",
            "Motor expenses",
            "Repairs and renewals",
            "Computer costs",
            "Sundry expenses",
            "Accountancy",
            "Commissions paid",
            "Entertainment",
            "Bank charges",
        )
    )
    depn = (
        nvl(rows, "Plant and machinery")
        + sum(r["dr"] for r in rows if r["row"] in (73, 74, 75, 76, 77))
    )
    # Use explicit P&L depn lines by row names that are FA labels
    depn_pl = 0.0
    for r in rows:
        if r["row"] >= 72 and r["row"] <= 77:
            depn_pl += r["dr"]
    depn_pl = round(depn_pl, 2)
    interest = nvl(rows, "Bank loan interest") + nvl(rows, "Hire purchase interest")
    disposal_gain = nvl(rows, "Loss on sale of motor vehicle", side="cr")

    cost_of_sales = round(purchases + haulage, 2)
    gross = round(sales - cost_of_sales, 2)
    staff = round(kelly_pay + swift_pay + wages + nic, 2)
    overheads = round(insurance + admin + staff + depn_pl, 2)
    op_profit = round(gross - overheads + disposal_gain, 2)
    profit_before = round(op_profit - interest, 2)

    fa = [
        ("Plant & machinery (owned)", plant_cost, plant_dep),
        ("Plant & machinery (leased / HP)", lease_p_cost, lease_p_dep),
        ("Fixtures & fittings", ff_cost, ff_dep),
        ("Motor vehicles (owned)", mv_cost, mv_dep),
        ("Motor vehicles (leased / HP)", lease_mv_cost, lease_mv_dep),
        ("Computer equipment", it_cost, it_dep),
    ]
    fa_cost = round(sum(c for _, c, _ in fa), 2)
    fa_dep = round(sum(d for _, _, d in fa), 2)
    fa_nbv = round(fa_cost - fa_dep, 2)

    ca = round(debtors + bank + vat, 2)
    creditors = round(
        kelly_cr + swift_cr + paye + accruals + oth_cr + loan + hp + dtax, 2
    )
    net_assets = round(fa_nbv + ca - creditors, 2)
    profit_for_year = profit_before  # no CT computed — deferred tax already on BS
    pl_cf = round(pl_bf + profit_for_year - div, 2)
    equity = round(shares + pl_cf, 2)

    return {
        "rows": rows,
        "tot_dr": tot_dr,
        "tot_cr": tot_cr,
        "fa": fa,
        "fa_cost": fa_cost,
        "fa_dep": fa_dep,
        "fa_nbv": fa_nbv,
        "debtors": debtors,
        "bank": bank,
        "vat": vat,
        "ca": ca,
        "kelly_cr": kelly_cr,
        "swift_cr": swift_cr,
        "paye": paye,
        "accruals": accruals,
        "oth_cr": oth_cr,
        "loan": loan,
        "hp": hp,
        "dtax": dtax,
        "creditors": creditors,
        "shares": shares,
        "pl_bf": pl_bf,
        "div": div,
        "sales": sales,
        "purchases": purchases,
        "haulage": haulage,
        "cos": cost_of_sales,
        "gross": gross,
        "insurance": insurance,
        "staff": staff,
        "kelly_pay": kelly_pay,
        "swift_pay": swift_pay,
        "wages": wages,
        "nic": nic,
        "admin": admin,
        "depn_pl": depn_pl,
        "overheads": overheads,
        "disposal_gain": disposal_gain,
        "op_profit": op_profit,
        "interest": interest,
        "profit": profit_before,
        "pl_cf": pl_cf,
        "net_assets": net_assets,
        "equity": equity,
        "entertainment": nvl(rows, "Entertainment"),
    }


def style_header(ws, row, cols, title=None):
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(color="FFFFFF", bold=True, name="Calibri")
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
    if title:
        ws.cell(row, 1, title)


def write_money(ws, r, c, value):
    cell = ws.cell(r, c, round(float(value or 0), 2))
    cell.number_format = '#,##0.00;(#,##0.00);"—"'
    cell.alignment = Alignment(horizontal="right")
    return cell


def write_pack(d: dict, dest: Path, iris_csv: Path) -> None:
    wb = Workbook()
    # --- Cover
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Go Green Pelleting Solutions Limited"
    cover["A1"].font = Font(name="Calibri", bold=True, size=18, color=NAVY)
    cover["A2"] = "Draft accounts working pack — year ended 31 March 2026"
    cover["A3"] = "Prepared from Mike Jacques trial balance (Outlook Holding, 23 July 2026)."
    cover["A4"] = "NOT an IRIS statutory set. Recode + import the IRIS Import sheet, then file in IRIS after client approval."
    cover["A6"] = "Supplied TB"
    cover["B6"] = "Balances"
    cover["A7"] = "Debits"
    write_money(cover, 7, 2, d["tot_dr"])
    cover["A8"] = "Credits"
    write_money(cover, 8, 2, d["tot_cr"])
    cover["A10"] = "Draft profit before tax"
    write_money(cover, 10, 2, d["profit"])
    cover["A11"] = "Dividends"
    write_money(cover, 11, 2, d["div"])
    cover["A12"] = "Net assets"
    write_money(cover, 12, 2, d["net_assets"])
    cover["A13"] = "Equity check (should equal net assets)"
    write_money(cover, 13, 2, d["equity"])
    cover["A15"] = "Open questions for Mike / you"
    cover["A16"] = (
        "1. 'Loss on sale of motor vehicle' is a CREDIT of £10,101.50 — treated as profit on disposal; proceeds implied £29,000."
    )
    cover["A17"] = "2. Two HP creditor lines (£106,298.01 and £145,887.26). Split <1 year / >1 year not given."
    cover["A18"] = "3. VAT £40,075.15 debtor — confirm repayable, not a mispost."
    cover["A19"] = "4. Entertainment £1,164.73 — add back for CT unless wholly staff."
    cover["A20"] = "5. No corporation tax charge computed (need capital allowances / losses). Deferred tax £50,962 left as on TB."
    cover["A21"] = "6. D Kelly / M Swift: £12,000 each in P&L (directors) and small credit balances on BS."
    cover.column_dimensions["A"].width = 92
    cover.column_dimensions["B"].width = 16

    # --- TB
    tb = wb.create_sheet("Supplied TB")
    tb["A1"] = "Account"
    tb["B1"] = "Debit"
    tb["C1"] = "Credit"
    style_header(tb, 1, 3)
    r = 2
    for row in d["rows"]:
        tb.cell(r, 1, row["name"])
        write_money(tb, r, 2, row["dr"])
        write_money(tb, r, 3, row["cr"])
        r += 1
    tb.cell(r, 1, "Total")
    tb.cell(r, 1).font = Font(bold=True)
    write_money(tb, r, 2, d["tot_dr"])
    write_money(tb, r, 3, d["tot_cr"])
    tb.column_dimensions["A"].width = 62
    tb.column_dimensions["B"].width = 16
    tb.column_dimensions["C"].width = 16

    # --- P&L
    pl = wb.create_sheet("Draft P&L")
    pl["A1"] = "Profit and loss account — year ended 31 March 2026"
    pl["A1"].font = Font(bold=True, size=14, color=NAVY)
    items = [
        ("Turnover", d["sales"]),
        ("Purchases", -d["purchases"]),
        ("Haulage", -d["haulage"]),
        ("Cost of sales", -d["cos"]),
        ("Gross profit", d["gross"]),
        ("Directors' remuneration (D Kelly / M Swift)", -d["kelly_pay"] - d["swift_pay"]),
        ("Wages", -d["wages"]),
        ("Social security", -d["nic"]),
        ("Insurance", -d["insurance"]),
        ("Administrative expenses", -d["admin"]),
        ("Depreciation", -d["depn_pl"]),
        ("Profit on disposal of motor vehicle", d["disposal_gain"]),
        ("Operating profit", d["op_profit"]),
        ("Interest payable", -d["interest"]),
        ("Profit before tax", d["profit"]),
        ("Tax (not computed — see Cover)", 0),
        ("Profit for the year (before tax)", d["profit"]),
    ]
    pl["A3"] = "Line"
    pl["B3"] = "£"
    style_header(pl, 3, 2)
    for i, (label, val) in enumerate(items, start=4):
        pl.cell(i, 1, label)
        write_money(pl, i, 2, val)
        if label in (
            "Gross profit",
            "Operating profit",
            "Profit before tax",
            "Profit for the year (before tax)",
        ):
            pl.cell(i, 1).font = Font(bold=True)
    pl.column_dimensions["A"].width = 48
    pl.column_dimensions["B"].width = 16

    # --- BS
    bs = wb.create_sheet("Draft balance sheet")
    bs["A1"] = "Balance sheet — 31 March 2026"
    bs["A1"].font = Font(bold=True, size=14, color=NAVY)
    bs["A3"] = "Fixed assets"
    bs["B3"] = "Cost"
    bs["C3"] = "Depreciation"
    bs["D3"] = "NBV"
    style_header(bs, 3, 4)
    r = 4
    for name, cost, dep in d["fa"]:
        bs.cell(r, 1, name)
        write_money(bs, r, 2, cost)
        write_money(bs, r, 3, dep)
        write_money(bs, r, 4, cost - dep)
        r += 1
    bs.cell(r, 1, "Total fixed assets")
    bs.cell(r, 1).font = Font(bold=True)
    write_money(bs, r, 2, d["fa_cost"])
    write_money(bs, r, 3, d["fa_dep"])
    write_money(bs, r, 4, d["fa_nbv"])
    r += 2
    bs.cell(r, 1, "Current assets")
    r += 1
    for label, val in (
        ("Other debtors", d["debtors"]),
        ("VAT repayable", d["vat"]),
        ("Cash at bank", d["bank"]),
        ("Total current assets", d["ca"]),
    ):
        bs.cell(r, 1, label)
        write_money(bs, r, 4, val)
        r += 1
    r += 1
    bs.cell(r, 1, "Creditors")
    r += 1
    for label, val in (
        ("D Kelly", d["kelly_cr"]),
        ("M Swift", d["swift_cr"]),
        ("PAYE", d["paye"]),
        ("Accruals", d["accruals"]),
        ("Other creditors", d["oth_cr"]),
        ("Bank loan", d["loan"]),
        ("Hire purchase (two TB lines combined)", d["hp"]),
        ("Deferred tax", d["dtax"]),
        ("Total creditors", d["creditors"]),
    ):
        bs.cell(r, 1, label)
        write_money(bs, r, 4, val)
        r += 1
    r += 1
    bs.cell(r, 1, "Net assets")
    bs.cell(r, 1).font = Font(bold=True)
    write_money(bs, r, 4, d["net_assets"])
    r += 2
    bs.cell(r, 1, "Capital and reserves")
    r += 1
    for label, val in (
        ("Called up share capital", d["shares"]),
        ("Profit and loss brought forward", d["pl_bf"]),
        ("Profit for the year (before tax)", d["profit"]),
        ("Dividends", -d["div"]),
        ("Profit and loss carried forward", d["pl_cf"]),
        ("Shareholders' funds", d["equity"]),
    ):
        bs.cell(r, 1, label)
        write_money(bs, r, 4, val)
        r += 1
    for col, w in (("A", 48), ("B", 16), ("C", 16), ("D", 16)):
        bs.column_dimensions[col].width = w

    # Line-for-line recode of Mike's TB (keeps the 1,705,293.71 balance).
    MAP = {
        "Plant and machinery cost brought forward": ("2100", "Plant and machinery cost"),
        "Plant and machinery additions": ("2100", "Plant and machinery cost"),
        "Plant and machinery transfer to ownership cost": ("2100", "Plant and machinery cost"),
        "Plant and machinery depreciation brought forward": ("2110", "Plant and machinery depreciation"),
        "Plant and machinery transfer to ownership depreciation": ("2110", "Plant and machinery depreciation"),
        "Plant and machinery charge for the year": ("2110", "Plant and machinery depreciation"),
        "Plant and machinery lease cost brought forward": ("2150", "Plant under lease / HP cost"),
        "Plant and machinery lease cost additions": ("2150", "Plant under lease / HP cost"),
        "Plant and machinery lease depreciation brought forward": ("2160", "Plant under lease / HP depn"),
        "Plant and mnachinery lease depreciation charge for the year": ("2160", "Plant under lease / HP depn"),
        "Fixtures and fittings cost brought forward": ("2200", "Fixtures and fittings cost"),
        "Fixtures and fittings depreciation brought forward": ("2210", "Fixtures and fittings depreciation"),
        "Fixtures and fittings depreciation charge for the year": ("2210", "Fixtures and fittings depreciation"),
        "Motor vehicles transfer to ownership cost": ("2300", "Motor vehicles cost"),
        "Motor vehicles disposal cost": ("2300", "Motor vehicles cost"),
        "Motor vehicles transfer to ownership depreciation": ("2310", "Motor vehicles depreciation"),
        "Motor vehicles depreciation on disposal": ("2310", "Motor vehicles depreciation"),
        "Motor vehicles depreciation charge for the year": ("2310", "Motor vehicles depreciation"),
        "Motor vehicles lease cost brought forward": ("2350", "Motor vehicles under lease cost"),
        "Motor vehicles lease cost transfer to ownership": ("2350", "Motor vehicles under lease cost"),
        "Motor vehicles lease depreciation brought forward": ("2360", "Motor vehicles under lease depn"),
        "Motor vehicles lease depreciation charge for the year": ("2360", "Motor vehicles under lease depn"),
        "Computer equipment cost brought forward": ("2400", "Computer equipment cost"),
        "Computer equipmemt depreciation brought forward": ("2410", "Computer equipment depreciation"),
        "Computer equipmemt depreciation charge for the year": ("2410", "Computer equipment depreciation"),
        "Other debtors": ("7105", "Other debtors"),
        "Bank account": ("7800", "Bank current account"),
        "D Kelly": ("8210", "D Kelly"),
        "M Swift": ("8215", "M Swift"),
        "VAT": ("7605", "VAT"),
        "PAYE": ("8230", "PAYE/NIC"),
        "Accruals": ("8300", "Accruals"),
        "Other creditors": ("8200", "Other creditors"),
        "Bank loan": ("8400", "Bank loan"),
        "HP": ("8450", "Hire purchase"),
        "Deferred tax": ("9100", "Deferred tax"),
        "Share capital": ("0010", "Called up share capital"),
        "Profit and loss account": ("0015", "Profit and loss account"),
        "Sales": ("4000", "Sales"),
        "Purchases": ("5000", "Purchases"),
        "Haulage": ("5200", "Haulage"),
        "Insurance": ("7100", "Insurance"),
        "Wages": ("7000", "Wages and salaries"),
        "Social security": ("7006", "Employers NIC"),
        "Telephone": ("7200", "Telephone"),
        "Post and stationery": ("7205", "Post and stationery"),
        "Travelling": ("7210", "Travelling"),
        "Motor expenses": ("7215", "Motor expenses"),
        "Repairs and renewals": ("7220", "Repairs and renewals"),
        "Computer costs": ("7225", "Computer costs"),
        "Sundry expenses": ("7230", "Sundry expenses"),
        "Accountancy": ("7300", "Accountancy"),
        "Commissions paid": ("7310", "Commissions paid"),
        "Entertainment": ("7400", "Entertainment"),
        "Bank charges": ("7500", "Bank charges"),
        "Plant and machinery": ("7600", "Depreciation"),
        "Fixtures and fittings": ("7600", "Depreciation"),
        "Motor vehicles": ("7600", "Depreciation"),
        "Computer equipment": ("7600", "Depreciation"),
        "Bank loan interest": ("7900", "Bank loan interest"),
        "Hire purchase interest": ("7910", "Hire purchase interest"),
        "Loss on sale of plant and equipment": ("7700", "Profit/loss on disposal"),
        "Loss on sale of motor vehicle": ("7700", "Profit/loss on disposal"),
        "Dividends": ("8000", "Dividends"),
    }
    BY_ROW = {
        14: ("2100", "Plant and machinery cost"),
        15: ("2110", "Plant and machinery depreciation"),
        19: ("2150", "Plant under lease / HP cost"),
        20: ("2160", "Plant under lease / HP depn"),
        26: ("2300", "Motor vehicles cost"),
        28: ("2310", "Motor vehicles depreciation"),
        32: ("2350", "Motor vehicles under lease cost"),
        34: ("2360", "Motor vehicles under lease depn"),
        57: ("6200", "Directors' remuneration"),
        58: ("6200", "Directors' remuneration"),
        41: ("8210", "D Kelly"),
        42: ("8215", "M Swift"),
        72: ("7600", "Depreciation"),
        73: ("7600", "Depreciation"),
        74: ("7600", "Depreciation"),
        75: ("7600", "Depreciation"),
        76: ("7600", "Depreciation"),
        77: ("7600", "Depreciation"),
    }
    recode_map = {}
    for row in d["rows"]:
        if row["row"] in BY_ROW:
            code, iname = BY_ROW[row["row"]]
        elif row["name"] == "D Kelly" and row["dr"]:
            code, iname = ("6200", "Directors' remuneration")
        elif row["name"] == "M Swift" and row["dr"]:
            code, iname = ("6200", "Directors' remuneration")
        elif row["name"] == "D Kelly":
            code, iname = ("8210", "D Kelly")
        elif row["name"] == "M Swift":
            code, iname = ("8215", "M Swift")
        else:
            code, iname = MAP.get(row["name"], ("9999", row["name"]))
        key = (code, iname)
        recode_map.setdefault(key, [0.0, 0.0])
        recode_map[key][0] += row["dr"]
        recode_map[key][1] += row["cr"]
    recode = [(c, n, round(dr, 2), round(cr, 2)) for (c, n), (dr, cr) in sorted(recode_map.items())]

    iris = wb.create_sheet("IRIS recode TB")
    iris["A1"] = "Suggested IRIS Elements nominals — check against this client's IRIS chart before import"
    iris["A3"] = "Code"
    iris["B3"] = "Name"
    iris["C3"] = "Debit"
    iris["D3"] = "Credit"
    style_header(iris, 3, 4)
    r = 4
    sdr = scr = 0.0
    for code, name, dr, cr in recode:
        if abs(dr) < 0.005 and abs(cr) < 0.005:
            continue
        iris.cell(r, 1, code)
        iris.cell(r, 2, name)
        write_money(iris, r, 3, dr)
        write_money(iris, r, 4, cr)
        sdr += dr
        scr += cr
        r += 1
    iris.cell(r, 2, "Total")
    write_money(iris, r, 3, sdr)
    write_money(iris, r, 4, scr)
    iris.column_dimensions["A"].width = 10
    iris.column_dimensions["B"].width = 44
    iris.column_dimensions["C"].width = 16
    iris.column_dimensions["D"].width = 16

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)

    iris_csv.parent.mkdir(parents=True, exist_ok=True)
    lines = ["Code,Name,Debit,Credit"]
    for code, name, dr, cr in recode:
        if abs(dr) < 0.005 and abs(cr) < 0.005:
            continue
        lines.append(f"{code},{name},{dr:.2f},{cr:.2f}")
    iris_csv.write_text("\n".join(lines) + "\n", encoding="utf-8")

    return sdr, scr


def mark_mike_clients() -> int:
    db = SessionLocal()
    n = 0
    try:
        for cid in MIKE_IDS:
            client = db.query(Client).filter(Client.id == cid).first()
            if not client:
                continue
            set_bookkeeping_source(db, client, "client_tb", note=NOTE)
            n += 1
    finally:
        db.close()
    return n


def main() -> int:
    n = mark_mike_clients()
    print(f"Mike Jacques clients set to client_tb: {n}")
    rows, tot_dr, tot_cr = parse_tb(SRC)
    d = build_pack(rows, tot_dr, tot_cr)
    dest = (
        CLIENT_DIR
        / "Current"
        / "Working Papers"
        / "Go Green Pelleting Solutions Limited - 31 March 2026 - Draft accounts pack.xlsx"
    )
    iris = (
        CLIENT_DIR
        / "Current"
        / "IRIS Import"
        / "2026-03-31 IRIS recode trial balance.csv"
    )
    sdr, scr = write_pack(d, dest, iris)
    print(f"TB {d['tot_dr']:.2f} / {d['tot_cr']:.2f}")
    print(f"PBT {d['profit']:.2f}  dividends {d['div']:.2f}  net assets {d['net_assets']:.2f}  equity {d['equity']:.2f}")
    print(f"IRIS recode {sdr:.2f} / {scr:.2f}")
    print(dest)
    print(iris)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
