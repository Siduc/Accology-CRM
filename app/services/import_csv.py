"""CSV/TSV import helpers with header detection and duplicate protection."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from app.models import Client, Person, Job
from app.services.dates import calculate_dates
from app.services.individuals import ensure_individual_client
from datetime import datetime, date


from app.services.company_numbers import normalize_company_number  # re-export

def excel_bytes_to_csv_text(content: bytes, sheet_name: Optional[str] = None) -> str:
    """Convert first (or named) sheet of an .xlsx file to CSV text."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel support requires openpyxl. Run: pip install openpyxl"
        ) from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    output = io.StringIO()
    writer = csv.writer(output)
    for row in ws.iter_rows(values_only=True):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        cells = []
        for cell in row:
            if cell is None:
                cells.append("")
            elif isinstance(cell, float) and cell.is_integer():
                cells.append(str(int(cell)))
            else:
                cells.append(str(cell).strip())
        writer.writerow(cells)
    wb.close()
    return output.getvalue()


def _detect_delimiter(text: str) -> str:
    sample = text[:4096]
    if "\t" in sample and sample.count("\t") >= sample.count(","):
        return "\t"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        return dialect.delimiter
    except csv.Error:
        return ","


def _normalize_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", h.strip().lower()).strip("_")


CLIENT_HEADER_MAP = {
    "company_name": "company_name",
    "company": "company_name",
    "name": "company_name",
    "client_name": "company_name",
    "client": "company_name",
    "organisation": "company_name",
    "organization": "company_name",
    "business_name": "company_name",
    "company_number": "company_number",
    "company_no": "company_number",
    "company_no_": "company_number",
    "companynumber": "company_number",
    "ch_number": "company_number",
    "ch_no": "company_number",
    "registration_number": "company_number",
    "reg_number": "company_number",
    "registered_number": "company_number",
    "co_number": "company_number",
    "contact_name": "contact_name",
    "contact": "contact_name",
    "primary_contact": "contact_name",
    "main_contact": "contact_name",
    "email": "email",
    "e_mail": "email",
    "email_address": "email",
    "phone": "phone",
    "telephone": "phone",
    "tel": "phone",
    "mobile": "phone",
    "address_line1": "address_line1",
    "address1": "address_line1",
    "address": "address_line1",
    "registered_office": "address_line1",
    "address_line2": "address_line2",
    "address2": "address_line2",
    "town": "town",
    "city": "town",
    "postcode": "postcode",
    "postal_code": "postcode",
    "client_type": "client_type",
    "type": "client_type",
    "entity_type": "client_type",
    "overall_status": "overall_status",
    "status": "overall_status",
    "client_status": "overall_status",
    "vat_number": "vat_number",
    "vat": "vat_number",
    "vat_no": "vat_number",
    "utr": "utr",
    "tax_reference": "utr",
    "notes": "notes",
    "note": "notes",
    "comments": "notes",
}

PERSON_HEADER_MAP = {
    "full_name": "full_name",
    "name": "full_name",
    "person_name": "full_name",
    "contact_name": "full_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "forename": "first_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "surname": "last_name",
    "email": "email",
    "e_mail": "email",
    "email_address": "email",
    "phone": "phone",
    "telephone": "phone",
    "tel": "phone",
    "mobile": "phone",
    "role": "role",
    "position": "role",
    "job_title": "role",
    "title": "role",
    "utr": "utr",
    "ni_number": "ni_number",
    "ni": "ni_number",
    "nino": "ni_number",
    "ch_code": "ch_code",
    "person_status": "person_status",
    "status": "person_status",
    "notes": "notes",
    "client_id": "client_id",
    "company_number": "company_number",
    "company_no": "company_number",
    "company": "company_name_link",
    "company_name": "company_name_link",
    "client": "company_name_link",
    "client_name": "company_name_link",
    "is_individual_client": "is_individual_client",
    "individual_client": "is_individual_client",
    "individual": "is_individual_client",
    "sa_client": "is_individual_client",
    "tax_only": "is_individual_client",
}

JOB_HEADER_MAP = {
    "title": "title",
    "type": "type",
    "job_type": "type",
    "client_id": "client_id",
    "company_number": "company_number",
    "period_end": "period_end",
    "fee": "fee",
    "status": "status",
    "is_recurring": "is_recurring",
    "recurring": "is_recurring",
    "notes": "notes",
}

CLIENT_POSITIONAL = [
    "company_name",
    "company_number",
    "contact_name",
    "email",
    "phone",
    "address_line1",
    "address_line2",
    "town",
    "postcode",
    "client_type",
    "overall_status",
    "vat_number",
    "utr",
    "notes",
]


@dataclass
class ImportResult:
    imported: int = 0
    skipped: int = 0
    errors: List[str] = field(default_factory=list)
    messages: List[str] = field(default_factory=list)


def parse_rows(text: str) -> Tuple[List[Dict[str, str]], List[str]]:
    """Parse CSV/TSV text into list of dicts. Returns (rows, warnings)."""
    warnings: List[str] = []
    if not text or not text.strip():
        return [], ["No data provided"]

    # Normalize line endings; strip BOM
    text = text.lstrip("\ufeff").replace("\r\n", "\n").replace("\r", "\n")
    delimiter = _detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    raw_rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not raw_rows:
        return [], ["No data rows found"]

    first = [_normalize_header(c) for c in raw_rows[0]]
    # Header if first row looks like known field names
    known = set(CLIENT_HEADER_MAP) | set(PERSON_HEADER_MAP) | set(JOB_HEADER_MAP)
    has_header = sum(1 for h in first if h in known) >= 2

    if has_header:
        headers = first
        data = raw_rows[1:]
    else:
        headers = []
        data = raw_rows
        warnings.append("No header detected — using positional column mapping")

    rows: List[Dict[str, str]] = []
    for row in data:
        if headers:
            mapped: Dict[str, str] = {}
            for i, cell in enumerate(row):
                if i >= len(headers):
                    break
                key = headers[i]
                mapped[key] = cell.strip()
            rows.append(mapped)
        else:
            mapped = {}
            for i, cell in enumerate(row):
                if i < len(CLIENT_POSITIONAL):
                    mapped[CLIENT_POSITIONAL[i]] = cell.strip()
                elif i == len(CLIENT_POSITIONAL):
                    mapped["notes"] = cell.strip()
                else:
                    mapped["notes"] = (mapped.get("notes") or "") + " " + cell.strip()
            rows.append(mapped)

    return rows, warnings


def _map_client_fields(row: Dict[str, str]) -> Dict[str, Optional[str]]:
    out: Dict[str, Optional[str]] = {}
    for raw_key, value in row.items():
        key = CLIENT_HEADER_MAP.get(raw_key, raw_key if raw_key in CLIENT_POSITIONAL else None)
        if key:
            out[key] = value if value else None
    # If company_number missing, try to find an 8-digit-ish token
    if not out.get("company_number"):
        for v in row.values():
            if v and re.fullmatch(r"[A-Z0-9]{6,8}", re.sub(r"\s+", "", v.upper())):
                out["company_number"] = normalize_company_number(v)
                break
    if out.get("company_number"):
        out["company_number"] = normalize_company_number(out["company_number"])
    return out


def import_clients(db: Session, text: str) -> ImportResult:
    result = ImportResult()
    rows, warnings = parse_rows(text)
    result.messages.extend(warnings)

    for idx, row in enumerate(rows, start=1):
        fields = _map_client_fields(row)
        company_number = fields.get("company_number")
        if not company_number:
            result.skipped += 1
            result.errors.append(f"Row {idx}: missing company_number")
            continue

        existing = (
            db.query(Client)
            .filter(Client.company_number == company_number)
            .first()
        )
        if existing:
            result.skipped += 1
            continue

        client = Client(
            company_name=fields.get("company_name"),
            company_number=company_number,
            contact_name=fields.get("contact_name"),
            email=fields.get("email"),
            phone=fields.get("phone"),
            address_line1=fields.get("address_line1"),
            address_line2=fields.get("address_line2"),
            town=fields.get("town"),
            postcode=fields.get("postcode"),
            client_type=fields.get("client_type"),
            overall_status=fields.get("overall_status") or "Active",
            vat_number=fields.get("vat_number"),
            utr=fields.get("utr"),
            notes=fields.get("notes"),
            source="csv",
        )
        db.add(client)
        result.imported += 1

    db.commit()
    result.messages.append(
        f"Imported {result.imported} clients, skipped {result.skipped}."
    )
    return result


def import_people(db: Session, text: str) -> ImportResult:
    result = ImportResult()
    rows, warnings = parse_rows(text)
    result.messages.extend(warnings)

    for idx, row in enumerate(rows, start=1):
        data: Dict[str, Optional[str]] = {}
        for raw_key, value in row.items():
            key = PERSON_HEADER_MAP.get(raw_key)
            if key:
                data[key] = value if value else None

        full_name = data.get("full_name")
        if not full_name:
            first = (data.get("first_name") or "").strip()
            last = (data.get("last_name") or "").strip()
            if first or last:
                full_name = f"{first} {last}".strip()
        if not full_name:
            # positional fallback
            vals = list(row.values())
            full_name = vals[0] if vals else None
            if not data.get("email") and len(vals) > 1:
                data["email"] = vals[1]
            if not data.get("phone") and len(vals) > 2:
                data["phone"] = vals[2]
            if not data.get("role") and len(vals) > 3:
                data["role"] = vals[3]

        if not full_name:
            result.skipped += 1
            result.errors.append(f"Row {idx}: missing full_name")
            continue

        client = None
        if data.get("client_id"):
            try:
                client = db.query(Client).filter(Client.id == int(data["client_id"])).first()
            except ValueError:
                pass
        if not client and data.get("company_number"):
            cn = normalize_company_number(data["company_number"])
            client = db.query(Client).filter(Client.company_number == cn).first()
        if not client and data.get("company_name_link"):
            name = data["company_name_link"].strip()
            client = (
                db.query(Client)
                .filter(Client.company_name.ilike(name))
                .first()
            )

        individual_flag = (data.get("is_individual_client") or "").strip().lower() in (
            "1",
            "y",
            "yes",
            "true",
            "individual",
            "sa",
            "tax only",
            "tax_only",
        )

        # Reuse existing person by name so one person can gain many companies
        person = (
            db.query(Person)
            .filter(Person.full_name.ilike(full_name.strip()))
            .first()
        )
        if person:
            changed = False
            if client and client not in person.clients:
                person.clients.append(client)
                changed = True
            if individual_flag and not person.is_individual_client:
                person.is_individual_client = True
                ensure_individual_client(db, person)
                changed = True
            if changed:
                result.imported += 1
            else:
                result.skipped += 1
            continue

        person = Person(
            full_name=full_name,
            email=data.get("email"),
            phone=data.get("phone"),
            role=data.get("role"),
            utr=data.get("utr"),
            ni_number=data.get("ni_number"),
            ch_code=data.get("ch_code"),
            person_status=(
                "Individual Client"
                if individual_flag
                else (data.get("person_status") or "Contact")
            ),
            notes=data.get("notes"),
            is_individual_client=individual_flag,
        )
        if client:
            person.clients.append(client)
        db.add(person)
        db.flush()
        if individual_flag:
            ensure_individual_client(db, person)
        result.imported += 1

    db.commit()
    result.messages.append(
        f"Imported/linked {result.imported} people rows, skipped {result.skipped}."
    )
    return result


def _parse_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def import_jobs(db: Session, text: str) -> ImportResult:
    result = ImportResult()
    rows, warnings = parse_rows(text)
    result.messages.extend(warnings)

    for idx, row in enumerate(rows, start=1):
        data: Dict[str, Optional[str]] = {}
        for raw_key, value in row.items():
            key = JOB_HEADER_MAP.get(raw_key)
            if key:
                data[key] = value if value else None

        # Positional: title, type, company_number, period_end, fee, status
        if not data.get("title") and not data.get("type"):
            vals = list(row.values())
            if vals:
                data.setdefault("title", vals[0] if vals else None)
            if len(vals) > 1:
                data.setdefault("type", vals[1])
            if len(vals) > 2:
                data.setdefault("company_number", vals[2])
            if len(vals) > 3:
                data.setdefault("period_end", vals[3])
            if len(vals) > 4:
                data.setdefault("fee", vals[4])
            if len(vals) > 5:
                data.setdefault("status", vals[5])

        client_id = None
        if data.get("client_id"):
            try:
                client_id = int(data["client_id"])
            except ValueError:
                pass
        if not client_id and data.get("company_number"):
            cn = normalize_company_number(data["company_number"])
            client = db.query(Client).filter(Client.company_number == cn).first()
            if client:
                client_id = client.id

        if not client_id:
            result.skipped += 1
            result.errors.append(f"Row {idx}: could not resolve client")
            continue

        job_type = data.get("type") or "Other"
        period_end = _parse_date(data.get("period_end"))
        statutory, target_start, target_completion = calculate_dates(job_type, period_end)

        fee = 0.0
        if data.get("fee"):
            try:
                fee = float(str(data["fee"]).replace("£", "").replace(",", ""))
            except ValueError:
                fee = 0.0

        title = data.get("title") or f"{job_type} - {period_end or 'TBC'}"
        job = Job(
            title=title,
            type=job_type,
            client_id=client_id,
            period_end=period_end,
            statutory_due_date=statutory,
            target_start=target_start,
            target_completion=target_completion,
            fee=fee,
            status=data.get("status") or "Planned",
            is_recurring=data.get("is_recurring") or "Yes",
            notes=data.get("notes"),
        )
        db.add(job)
        result.imported += 1

    db.commit()
    result.messages.append(f"Imported {result.imported} jobs, skipped {result.skipped}.")
    return result
